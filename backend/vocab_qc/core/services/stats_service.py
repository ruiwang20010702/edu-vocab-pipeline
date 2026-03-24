"""仪表板统计服务."""

from datetime import UTC, datetime, timedelta

from sqlalchemy import func, literal_column
from sqlalchemy.orm import Session

from vocab_qc.core.cache import TTLCache
from vocab_qc.core.models.content_layer import ContentItem
from vocab_qc.core.models.data_layer import Word
from vocab_qc.core.models.enums import QcStatus
from vocab_qc.core.models.package_layer import Package
from vocab_qc.core.models.quality_layer import AiUsageLog, QcRuleResult

# 终态集合：approved 或 rejected
_TERMINAL_STATUSES = [QcStatus.APPROVED.value, QcStatus.REJECTED.value]

# Stats 缓存（10s TTL，减少前端轮询对数据库的压力）
stats_cache = TTLCache(default_ttl=10.0)
_STATS_CACHE_KEY = "dashboard_stats"


def invalidate_stats_cache() -> None:
    """主动失效 stats 缓存（导入/生产/审核等写操作后调用）。"""
    stats_cache.invalidate(_STATS_CACHE_KEY)


def get_dashboard_stats(session: Session) -> dict:
    """聚合统计：总词数、已入库、待处理、未通过、通过率、Bad Case 分类。

    定义对齐总表：
    - 已入库(approved_count) = 所有 ContentItem 均为终态的词数
    - 待处理(pending_count) = total_words - approved_count（存在非终态项的词数）
    - 未通过(rejected_count) = 存在 layer1/2_failed 项的词数（信息性统计）
    """
    cached = stats_cache.get(_STATS_CACHE_KEY)
    if cached is not None:
        return cached

    total_words = session.query(func.count()).select_from(Word).scalar() or 0

    # 已入库 = 有 ContentItem 且全部为终态的词数
    # NOT EXISTS 比 NOT IN 更高效（避免子查询物化 + NULL 安全）
    from sqlalchemy import alias

    ci1 = alias(ContentItem.__table__, name="ci1")
    ci2 = alias(ContentItem.__table__, name="ci2")

    # 有内容的 word_id（去重）
    has_content_q = session.query(ci1.c.word_id.distinct())
    # 存在非终态项的子查询
    non_terminal_exists = (
        session.query(ci2.c.id)
        .filter(
            ci2.c.word_id == ci1.c.word_id,
            ~ci2.c.qc_status.in_(_TERMINAL_STATUSES),
        )
        .exists()
    )
    approved_count = (
        has_content_q.filter(~non_terminal_exists).count()
    )

    pending_count = total_words - approved_count

    rejected_count = (
        session.query(func.count(func.distinct(ContentItem.word_id)))
        .filter(
            ContentItem.qc_status.in_([
                QcStatus.LAYER1_FAILED.value,
                QcStatus.LAYER2_FAILED.value,
            ])
        )
        .scalar()
        or 0
    )

    pass_rate = round(approved_count / total_words * 100, 1) if total_words > 0 else 0.0

    # Bad Case 分类：按 rule_id + dimension 聚合失败数（仅统计最新质检结果）
    issue_rows = (
        session.query(
            QcRuleResult.rule_id,
            QcRuleResult.dimension,
            func.count().label("count"),
        )
        .join(ContentItem, ContentItem.id == QcRuleResult.content_item_id)
        .filter(
            QcRuleResult.passed == False,  # noqa: E712
            QcRuleResult.run_id == ContentItem.last_qc_run_id,
        )
        .group_by(QcRuleResult.rule_id, QcRuleResult.dimension)
        .all()
    )
    issues = [
        {"field": row.rule_id, "dimension": row.dimension, "count": row.count}
        for row in issue_rows
    ]

    result = {
        "total_words": total_words,
        "approved_count": approved_count,
        "pending_count": pending_count,
        "rejected_count": rejected_count,
        "pass_rate": pass_rate,
        "issues": issues,
    }
    stats_cache.set(_STATS_CACHE_KEY, result)
    return result


def get_ai_usage_stats(session: Session, days: int = 7) -> dict:
    """AI 用量统计：总量 + 按维度×阶段分组 + 日趋势。"""
    cache_key = f"ai_usage_stats_{days}"
    cached = stats_cache.get(cache_key)
    if cached is not None:
        return cached

    since = datetime.now(UTC) - timedelta(days=days)

    # 聚合总量
    totals = session.query(
        func.coalesce(func.sum(AiUsageLog.total_tokens), 0),
        func.sum(AiUsageLog.estimated_cost_usd),
        func.count(),
    ).filter(AiUsageLog.created_at >= since).one()

    total_tokens = int(totals[0])
    total_cost = round(float(totals[1]), 6) if totals[1] is not None else None
    total_calls = int(totals[2])

    # 按维度 + 阶段分组
    by_dim_rows = (
        session.query(
            AiUsageLog.dimension,
            AiUsageLog.phase,
            func.coalesce(func.sum(AiUsageLog.total_tokens), 0).label("total_tokens"),
            func.sum(AiUsageLog.estimated_cost_usd).label("cost"),
            func.count().label("call_count"),
        )
        .filter(AiUsageLog.created_at >= since)
        .group_by(AiUsageLog.dimension, AiUsageLog.phase)
        .all()
    )
    by_dimension = [
        {
            "dimension": row.dimension,
            "phase": row.phase,
            "total_tokens": int(row.total_tokens),
            "estimated_cost_usd": round(float(row.cost), 6) if row.cost is not None else None,
            "call_count": int(row.call_count),
        }
        for row in by_dim_rows
    ]

    # 日趋势（兼容 PostgreSQL date_trunc 和 SQLite strftime）
    bind = session.bind
    dialect = bind.dialect.name if bind else "sqlite"
    if dialect == "postgresql":
        day_col = func.date_trunc(literal_column("'day'"), AiUsageLog.created_at).label("day")
    else:
        day_col = func.strftime("%Y-%m-%d", AiUsageLog.created_at).label("day")

    trend_rows = (
        session.query(
            day_col,
            func.coalesce(func.sum(AiUsageLog.total_tokens), 0).label("total_tokens"),
            func.count().label("call_count"),
        )
        .filter(AiUsageLog.created_at >= since)
        .group_by("day")
        .order_by("day")
        .all()
    )
    daily_trend = [
        {
            "date": str(row.day)[:10],
            "total_tokens": int(row.total_tokens),
            "call_count": int(row.call_count),
        }
        for row in trend_rows
    ]

    result = {
        "total_tokens": total_tokens,
        "total_cost_usd": total_cost,
        "total_calls": total_calls,
        "by_dimension": by_dimension,
        "daily_trend": daily_trend,
    }
    stats_cache.set(cache_key, result)
    return result


def get_production_records(session: Session) -> dict:
    """按 Package 聚合 AI 用量，返回生产记录数据。"""
    from sqlalchemy import case

    gemini_filter = func.lower(AiUsageLog.ai_model).like("%gemini%")
    gpt_filter = func.lower(AiUsageLog.ai_model).like("%gpt%")

    rows = (
        session.query(
            Package.id,
            Package.name,
            Package.total_words,
            Package.started_at,
            Package.completed_at,
            func.coalesce(
                func.sum(case((gemini_filter, AiUsageLog.prompt_tokens), else_=0)), 0,
            ).label("gemini_in"),
            func.coalesce(
                func.sum(case((gemini_filter, AiUsageLog.completion_tokens), else_=0)), 0,
            ).label("gemini_out"),
            func.coalesce(
                func.sum(case((gpt_filter, AiUsageLog.prompt_tokens), else_=0)), 0,
            ).label("gpt_in"),
            func.coalesce(
                func.sum(case((gpt_filter, AiUsageLog.completion_tokens), else_=0)), 0,
            ).label("gpt_out"),
        )
        .outerjoin(AiUsageLog, AiUsageLog.package_id == Package.id)
        .filter(Package.status.in_(["completed", "processing", "failed"]))
        .group_by(Package.id, Package.name, Package.total_words, Package.started_at, Package.completed_at)
        .order_by(Package.id)
        .all()
    )

    records = []
    totals = {"words": 0, "gemini_input": 0, "gemini_output": 0, "gpt_input": 0, "gpt_output": 0}

    for row in rows:
        rec = {
            "batch_id": row[0],
            "batch_name": row[1],
            "word_count": row[2],
            "started_at": row[3].isoformat() if row[3] else None,
            "completed_at": row[4].isoformat() if row[4] else None,
            "gemini_input_tokens": int(row[5]),
            "gemini_output_tokens": int(row[6]),
            "gpt_input_tokens": int(row[7]),
            "gpt_output_tokens": int(row[8]),
        }
        records.append(rec)
        totals["words"] += rec["word_count"]
        totals["gemini_input"] += rec["gemini_input_tokens"]
        totals["gemini_output"] += rec["gemini_output_tokens"]
        totals["gpt_input"] += rec["gpt_input_tokens"]
        totals["gpt_output"] += rec["gpt_output_tokens"]

    return {"records": records, "totals": totals}


def export_production_records_xlsx(session: Session):
    """导出生产记录为 xlsx（BytesIO），列结构与成本估算表一致。"""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    data = get_production_records(session)

    wb = Workbook()
    ws = wb.active
    ws.title = "生产记录"

    # 样式
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("批次号", 8),
        ("批次名称", 20),
        ("词数", 8),
        ("开始时间", 18),
        ("结束时间", 18),
        ("Gemini\nInput Tokens", 15),
        ("Gemini\nOutput Tokens", 15),
        ("GPT\nInput Tokens", 15),
        ("GPT\nOutput Tokens", 15),
        ("备注", 15),
    ]

    for col_idx, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    # 数据行
    for row_idx, rec in enumerate(data["records"], 2):
        ws.cell(row=row_idx, column=1, value=rec["batch_id"])
        ws.cell(row=row_idx, column=2, value=rec["batch_name"])
        ws.cell(row=row_idx, column=3, value=rec["word_count"])
        ws.cell(row=row_idx, column=4, value=rec["started_at"] or "")
        ws.cell(row=row_idx, column=5, value=rec["completed_at"] or "")
        ws.cell(row=row_idx, column=6, value=rec["gemini_input_tokens"])
        ws.cell(row=row_idx, column=7, value=rec["gemini_output_tokens"])
        ws.cell(row=row_idx, column=8, value=rec["gpt_input_tokens"])
        ws.cell(row=row_idx, column=9, value=rec["gpt_output_tokens"])

    # 汇总行
    if data["records"]:
        total_row = len(data["records"]) + 2
        ws.cell(row=total_row, column=1, value="汇总").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=data["totals"]["words"]).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=data["totals"]["gemini_input"]).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=data["totals"]["gemini_output"]).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=data["totals"]["gpt_input"]).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=data["totals"]["gpt_output"]).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
