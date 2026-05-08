"""数据导入服务: 解析文件 → 创建 Word / Meaning / ContentItem。"""

import csv
import io
import json
import logging
from typing import Any

_logger = logging.getLogger(__name__)

# 字段长度限制
_MAX_WORD_LEN = 100
_MAX_POS_LEN = 20
_MAX_DEFINITION_LEN = 500
_MAX_SOURCE_LEN = 200
_MAX_IPA_LEN = 200
_MAX_AUDIO_URL_LEN = 500

from sqlalchemy.orm import Session

from vocab_qc.core.models.content_layer import ContentItem
from vocab_qc.core.models.data_layer import Meaning, Phonetic, Source, Word
from vocab_qc.core.models.enums import QcStatus
from vocab_qc.core.models.package_layer import Package, PackageWord
from vocab_qc.core.models.quality_layer import RetryCounter, ReviewItem


def _truncate_field(value: str, max_len: int, field_name: str, word: str) -> str:
    """截断超长字段并记录警告。"""
    if len(value) > max_len:
        _logger.warning("字段 %s 超长截断 (len=%d>%d) word=%s", field_name, len(value), max_len, word)
        return value[:max_len]
    return value


def import_from_json(session: Session, data: list[dict[str, Any]], batch_name: str, *, force: bool = False, warnings: list[str] | None = None) -> dict:
    """从 JSON 数据导入词汇。

    期望格式:
    [
      {
        "word": "kind",
        "meanings": [
          {"pos": "adj.", "definition": "善良的", "sources": ["人教七上U1"]}
        ]
      }
    ]

    性能优化：预批量查询 Word/Meaning/Phonetic，延迟 flush，
    避免逐条查询导致的 N+1 问题。4400 词从 ~15 分钟降至 <10 秒。
    """
    package = _get_or_create_package(session, batch_name, force=force)
    all_warnings: list[str] = list(warnings) if warnings else []

    # ── 第一遍：收集所有待处理的 word_text 和 meaning 信息 ──
    parsed_entries: list[dict[str, Any]] = []
    all_word_texts: list[str] = []
    for entry in data:
        word_text = entry.get("word", "").strip()
        if not word_text:
            continue
        all_word_texts.append(word_text)
        parsed_entries.append(entry)

    if not all_word_texts:
        return {"batch_id": str(package.id), "word_count": 0, "warnings": all_warnings}

    # ── 预批量查询已存在的 Word ──
    unique_word_texts = list(dict.fromkeys(all_word_texts))  # 保序去重
    existing_words: dict[str, Word] = {}
    for i in range(0, len(unique_word_texts), 900):
        chunk = unique_word_texts[i:i + 900]
        for w in session.query(Word).filter(Word.word.in_(chunk)).all():
            existing_words[w.word] = w

    # ── 批量创建不存在的 Word（一次 flush） ──
    new_words: list[Word] = []
    for wt in unique_word_texts:
        if wt not in existing_words:
            w = Word(word=wt)
            session.add(w)
            new_words.append(w)
    if new_words:
        session.flush()  # 一次 flush 获取所有新 Word 的 id
        for w in new_words:
            existing_words[w.word] = w

    # ── 预批量查询已存在的 Phonetic ──
    all_word_ids = [existing_words[wt].id for wt in unique_word_texts]
    existing_phonetic_ids: set[int] = set()
    for i in range(0, len(all_word_ids), 900):
        chunk = all_word_ids[i:i + 900]
        existing_phonetic_ids.update(
            row[0] for row in session.query(Phonetic.word_id).filter(
                Phonetic.word_id.in_(chunk),
            ).all()
        )

    # ── 预批量查询已存在的 Meaning ──
    existing_meanings: dict[tuple[int, str, str], Meaning] = {}
    for i in range(0, len(all_word_ids), 900):
        chunk = all_word_ids[i:i + 900]
        for m in session.query(Meaning).filter(Meaning.word_id.in_(chunk)).all():
            existing_meanings[(m.word_id, m.pos, m.definition)] = m

    # ── 第二遍：构建所有对象，延迟 flush ──
    word_count = 0
    imported_words: list[Word] = []
    imported_meanings: list[tuple[Word, Meaning]] = []
    pending_sources: list[tuple[int, str, dict]] = []  # (meaning_id, source_name, extra) - 暂存
    pending_source_entries: list[tuple[str, str, str, dict]] = []  # (word_text, meaning_key, source_name, extra)
    new_phonetics: list[Phonetic] = []
    new_meanings: list[Meaning] = []
    # meaning_key → Meaning 对象（含尚未 flush 的新建对象）
    meaning_lookup: dict[tuple[int, str, str], Meaning] = dict(existing_meanings)

    for entry in parsed_entries:
        word_text = entry.get("word", "").strip()
        word = existing_words[word_text]
        word_count += 1
        imported_words.append(word)

        # 音标（IPA）— 批量处理（支持英式/美式分离）
        ipa_uk = entry.get("ipa_uk", "").strip()
        ipa_us = entry.get("ipa_us", "").strip()
        audio_url_uk = entry.get("audio_url_uk", "").strip()
        audio_url_us = entry.get("audio_url_us", "").strip()
        has_phonetic_data = ipa_uk or ipa_us or audio_url_uk or audio_url_us
        if has_phonetic_data:
            if word.id in existing_phonetic_ids:
                phonetic = session.query(Phonetic).filter_by(word_id=word.id).first()
                if phonetic:
                    if ipa_uk:
                        phonetic.ipa_uk = ipa_uk
                    if ipa_us:
                        phonetic.ipa_us = ipa_us
                    if audio_url_uk:
                        phonetic.audio_url_uk = audio_url_uk
                    if audio_url_us:
                        phonetic.audio_url_us = audio_url_us
            else:
                new_phonetics.append(Phonetic(
                    word_id=word.id,
                    ipa_uk=ipa_uk or None,
                    ipa_us=ipa_us or None,
                    audio_url_uk=audio_url_uk or None,
                    audio_url_us=audio_url_us or None,
                    syllables="",
                ))
                existing_phonetic_ids.add(word.id)

        for m_data in entry.get("meanings", []):
            pos = m_data.get("pos", "").strip()
            definition = m_data.get("definition", "").strip()
            if not pos or not definition:
                if not definition:
                    msg = f"单词 '{word_text}' 缺失释义，已跳过该义项"
                    if msg not in all_warnings:
                        all_warnings.append(msg)
                continue

            key = (word.id, pos, definition)
            meaning = meaning_lookup.get(key)
            if meaning is None:
                meaning = Meaning(word_id=word.id, pos=pos, definition=definition)
                session.add(meaning)
                new_meanings.append(meaning)
                meaning_lookup[key] = meaning

            imported_meanings.append((word, meaning))

            source_extra = {
                "textbook_id": entry.get("textbook_id"),
                "word_book_id": entry.get("word_book_id"),
                "unit_id": entry.get("unit_id"),
            }
            sources = m_data.get("sources", [])
            has_source_ids = any(v is not None for v in source_extra.values())
            m_key = f"{word.id}:{pos}:{definition}"
            if sources:
                for src_name in sources:
                    pending_source_entries.append((word_text, m_key, src_name, source_extra))
            elif has_source_ids:
                # 无 source 列但有教材 ID → 用批次名作 source_name
                pending_source_entries.append((word_text, m_key, batch_name, source_extra))

    # ── 统一 flush：新 Phonetic + 新 Meaning ──
    if new_phonetics:
        session.add_all(new_phonetics)
    if new_meanings or new_phonetics:
        session.flush()  # 获取所有新 Meaning 的 id

    # ── 批量处理 Source（Meaning id 已就绪） ──
    pending_sources = []
    for word_text, m_key, src_name, src_extra in pending_source_entries:
        parts = m_key.split(":", 2)
        key = (int(parts[0]), parts[1], parts[2])
        meaning = meaning_lookup.get(key)
        if meaning and meaning.id:
            pending_sources.append((meaning.id, src_name, src_extra))

    if pending_sources:
        # 去重 key = (meaning_id, source_name, textbook_id, word_book_id, unit_id)
        # 同一义项+同一来源名但不同教材/词书/单元 → 保留为不同记录
        existing_source_keys: set[tuple] = set()
        unique_meaning_ids = list({mid for mid, _, _ in pending_sources})
        for i in range(0, len(unique_meaning_ids), 900):
            chunk = unique_meaning_ids[i:i + 900]
            for row in session.query(
                Source.meaning_id, Source.source_name,
                Source.textbook_id, Source.word_book_id, Source.unit_id,
            ).filter(Source.meaning_id.in_(chunk)).all():
                existing_source_keys.add((row[0], row[1], row[2], row[3], row[4]))

        new_sources = []
        seen_keys: set[tuple] = set()
        for mid, sname, src_extra in pending_sources:
            key = (mid, sname, src_extra.get("textbook_id"), src_extra.get("word_book_id"), src_extra.get("unit_id"))
            if key not in existing_source_keys and key not in seen_keys:
                new_sources.append(Source(
                    meaning_id=mid,
                    source_name=sname,
                    textbook_id=src_extra.get("textbook_id"),
                    word_book_id=src_extra.get("word_book_id"),
                    unit_id=src_extra.get("unit_id"),
                ))
                seen_keys.add(key)
        if new_sources:
            session.add_all(new_sources)

    # ── 批量处理 PackageWord ──
    pending_word_ids = [existing_words[wt].id for wt in dict.fromkeys(all_word_texts)]
    if pending_word_ids:
        existing_pw_word_ids: set[int] = set()
        for i in range(0, len(pending_word_ids), 900):
            chunk = pending_word_ids[i:i + 900]
            existing_pw_word_ids.update(
                row[0] for row in session.query(PackageWord.word_id).filter(
                    PackageWord.package_id == package.id,
                    PackageWord.word_id.in_(chunk),
                ).all()
            )
        new_pws = []
        seen_word_ids: set[int] = set()
        for wid in pending_word_ids:
            if wid not in existing_pw_word_ids and wid not in seen_word_ids:
                new_pws.append(PackageWord(package_id=package.id, word_id=wid))
                seen_word_ids.add(wid)
        if new_pws:
            session.add_all(new_pws)

    session.flush()

    # ── 创建 ContentItem 占位记录（分批写入，避免大事务锁） ──
    _create_content_placeholders(session, imported_words, imported_meanings)

    # 更新 Package 统计
    package.status = "pending"
    package.total_words = word_count
    package.processed_words = 0

    session.flush()
    _logger.info(
        "导入完成 batch=%s 新建词=%d 复用词=%d 新义项=%d 新音标=%d 总词数=%d",
        batch_name, len(new_words), word_count - len(new_words),
        len(new_meanings), len(new_phonetics), word_count,
    )
    if all_warnings:
        for w in all_warnings:
            _logger.warning("导入警告 batch=%s: %s", batch_name, w)
    return {"batch_id": str(package.id), "word_count": word_count, "warnings": all_warnings}


def _parse_csv_text(text: str) -> tuple[list[dict[str, Any]], list[str]]:
    """将 CSV 文本解析为词汇 entries 列表，同时收集缺失释义的警告。"""
    reader = csv.DictReader(io.StringIO(text))
    entries: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    warned_words: set[str] = set()
    for row in reader:
        word = row.get("word", "").strip()
        if not word:
            continue
        if len(word) > _MAX_WORD_LEN:
            _logger.warning("CSV 跳过超长单词 (len=%d): %s...", len(word), word[:30])
            continue
        if word not in entries:
            entries[word] = {"word": word, "meanings": []}
        pos = _truncate_field(row.get("pos", "").strip(), _MAX_POS_LEN, "pos", word)
        definition = _truncate_field(row.get("definition", "").strip(), _MAX_DEFINITION_LEN, "definition", word)
        source = _truncate_field(row.get("source", "").strip(), _MAX_SOURCE_LEN, "source", word)
        ipa_uk = _truncate_field(row.get("ipa_uk", row.get("ipa", "")).strip(), _MAX_IPA_LEN, "ipa_uk", word)
        ipa_us = _truncate_field(row.get("ipa_us", "").strip(), _MAX_IPA_LEN, "ipa_us", word)
        if ipa_uk and not entries[word].get("ipa_uk"):
            entries[word]["ipa_uk"] = ipa_uk
        if ipa_us and not entries[word].get("ipa_us"):
            entries[word]["ipa_us"] = ipa_us
        # 教材结构化 ID
        tb = row.get("textbook_id", "").strip()
        wb = row.get("word_book_id", "").strip()
        ui = row.get("unit_id", "").strip()
        if tb and not entries[word].get("textbook_id"):
            entries[word]["textbook_id"] = tb
        if wb and not entries[word].get("word_book_id"):
            entries[word]["word_book_id"] = wb
        if ui and not entries[word].get("unit_id"):
            entries[word]["unit_id"] = ui
        if pos and definition:
            entries[word]["meanings"].append({
                "pos": pos,
                "definition": definition,
                "sources": [source] if source else [],
            })
        elif not definition and word not in warned_words:
            warnings.append(f"单词 '{word}' 缺失释义，已跳过该义项")
            warned_words.add(word)
    return list(entries.values()), warnings


def import_from_csv(session: Session, content: str, batch_name: str) -> dict:
    """从 CSV 导入。期望列: word, pos, definition, source。"""
    entries, warnings = _parse_csv_text(content)
    return import_from_json(session, entries, batch_name, warnings=warnings)


def _parse_excel(file_content: bytes) -> list[dict[str, Any]]:
    """将 Excel 文件解析为词汇 entries 列表。

    期望列: word, pos, definition, source（与 CSV 格式一致）。
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法解析 Excel 文件: {exc}") from exc
    ws = wb.active
    if ws is None:
        raise ValueError("Excel 文件中没有工作表")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Excel 文件为空")

    # 首行为表头，查找列索引
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map: dict[str, int] = {}
    for alias, key in [
        ("word", "word"), ("单词", "word"), ("单词名", "word"),
        ("pos", "pos"), ("词性", "pos"),
        ("definition", "definition"), ("释义", "definition"), ("中文释义", "definition"), ("剑桥释义", "definition"),
        ("source", "source"), ("来源", "source"), ("教材来源", "source"),
        ("ipa", "ipa_uk"), ("音标", "ipa_uk"), ("ipa_uk", "ipa_uk"), ("英式音标", "ipa_uk"),
        ("ipa_us", "ipa_us"), ("美式音标", "ipa_us"),
        ("audio_url_uk", "audio_url_uk"), ("英式音频url", "audio_url_uk"), ("英式发音url", "audio_url_uk"),
        ("audio_url_us", "audio_url_us"), ("美式音频url", "audio_url_us"), ("美式发音url", "audio_url_us"),
        ("textbook_id", "textbook_id"), ("教材id", "textbook_id"),
        ("word_book_id", "word_book_id"), ("词书id", "word_book_id"),
        ("unit_id", "unit_id"), ("单元id", "unit_id"),
    ]:
        if alias in header and key not in col_map:
            col_map[key] = header.index(alias)

    if "word" not in col_map:
        raise ValueError("Excel 缺少必要的 'word'（或 '单词'）列")

    entries: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    warned_words: set[str] = set()
    for row in rows[1:]:
        word = str(row[col_map["word"]] or "").strip()
        if not word:
            continue
        if len(word) > _MAX_WORD_LEN:
            _logger.warning("Excel 跳过超长单词 (len=%d): %s...", len(word), word[:30])
            continue
        if word not in entries:
            entries[word] = {"word": word, "meanings": []}
        pos = _truncate_field(
            str(row[col_map.get("pos", -1)] or "").strip() if "pos" in col_map else "",
            _MAX_POS_LEN, "pos", word,
        )
        definition = _truncate_field(
            str(row[col_map.get("definition", -1)] or "").strip() if "definition" in col_map else "",
            _MAX_DEFINITION_LEN, "definition", word,
        )
        source = _truncate_field(
            str(row[col_map.get("source", -1)] or "").strip() if "source" in col_map else "",
            _MAX_SOURCE_LEN, "source", word,
        )
        ipa_uk = _truncate_field(
            str(row[col_map.get("ipa_uk", -1)] or "").strip() if "ipa_uk" in col_map else "",
            _MAX_IPA_LEN, "ipa_uk", word,
        )
        ipa_us = _truncate_field(
            str(row[col_map.get("ipa_us", -1)] or "").strip() if "ipa_us" in col_map else "",
            _MAX_IPA_LEN, "ipa_us", word,
        )
        if ipa_uk and not entries[word].get("ipa_uk"):
            entries[word]["ipa_uk"] = ipa_uk
        if ipa_us and not entries[word].get("ipa_us"):
            entries[word]["ipa_us"] = ipa_us
        # 音频 URL
        audio_uk = _truncate_field(
            str(row[col_map.get("audio_url_uk", -1)] or "").strip() if "audio_url_uk" in col_map else "",
            _MAX_AUDIO_URL_LEN, "audio_url_uk", word,
        )
        audio_us = _truncate_field(
            str(row[col_map.get("audio_url_us", -1)] or "").strip() if "audio_url_us" in col_map else "",
            _MAX_AUDIO_URL_LEN, "audio_url_us", word,
        )
        if audio_uk and not entries[word].get("audio_url_uk"):
            entries[word]["audio_url_uk"] = audio_uk
        if audio_us and not entries[word].get("audio_url_us"):
            entries[word]["audio_url_us"] = audio_us
        # 教材结构化 ID
        def _safe_str(val: object) -> str | None:
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        if "textbook_id" in col_map and not entries[word].get("textbook_id"):
            entries[word]["textbook_id"] = _safe_str(row[col_map["textbook_id"]])
        if "word_book_id" in col_map and not entries[word].get("word_book_id"):
            entries[word]["word_book_id"] = _safe_str(row[col_map["word_book_id"]])
        if "unit_id" in col_map and not entries[word].get("unit_id"):
            entries[word]["unit_id"] = _safe_str(row[col_map["unit_id"]])
        if pos and definition:
            entries[word]["meanings"].append({
                "pos": pos,
                "definition": definition,
                "sources": [source] if source else [],
            })
        elif not definition and word not in warned_words:
            warnings.append(f"单词 '{word}' 缺失释义，已跳过该义项")
            warned_words.add(word)
    return list(entries.values()), warnings


# S-H2: Magic bytes 校验映射
_MAGIC_BYTES = {
    ".xlsx": [b"PK\x03\x04"],  # ZIP (OOXML)
    ".xls": [b"\xd0\xcf\x11\xe0"],  # OLE2
    ".json": [],  # JSON 靠解析校验
    ".csv": [],   # CSV 靠解析校验
}


def _validate_magic_bytes(file_content: bytes, filename: str) -> None:
    """校验文件 magic bytes 与扩展名一致，防止恶意文件伪装。"""
    lower = filename.lower()
    for ext, signatures in _MAGIC_BYTES.items():
        if lower.endswith(ext):
            if not signatures:
                return  # JSON/CSV 无固定 magic bytes
            for sig in signatures:
                if file_content[:len(sig)] == sig:
                    return
            raise ValueError(f"文件内容与 {ext} 格式不匹配，请确认文件完整性")
    raise ValueError(f"不支持的文件格式: {filename}")


def parse_upload(file_content: bytes, filename: str) -> tuple[list[dict[str, Any]], list[str]]:
    """根据文件扩展名解析上传内容，返回 (entries, warnings)。"""
    _validate_magic_bytes(file_content, filename)
    lower = filename.lower()
    if lower.endswith(".json"):
        data = json.loads(file_content.decode("utf-8"))
        if not isinstance(data, list):
            raise ValueError("JSON 文件格式错误：期望数组格式，如 [{\"word\": \"hello\", ...}]")
        return data, []
    if lower.endswith(".csv"):
        return _parse_csv_text(file_content.decode("utf-8"))
    if lower.endswith((".xlsx", ".xls")):
        return _parse_excel(file_content)
    raise ValueError(f"不支持的文件格式: {filename}，请使用 .xlsx, .csv 或 .json")


def _get_or_create_package(session: Session, name: str, *, force: bool = False) -> Package:
    pkg = session.query(Package).filter_by(name=name).first()
    if pkg is not None:
        if pkg.status in ("pending", "importing"):
            return pkg
        if not force:
            raise ValueError(f"批次 '{name}' 已在处理中（状态: {pkg.status}），不可重复导入")
        # force=True: 清理旧数据，重新导入
        _clean_package_data(session, pkg)
        pkg.status = "pending"
        pkg.processed_words = 0
        session.flush()
        return pkg
    pkg = Package(name=name)
    session.add(pkg)
    session.flush()
    return pkg


def _clean_package_data(session: Session, pkg: Package) -> None:
    """清理批次关联的旧数据，为重新导入做准备。

    策略：
    - 删除该包独占词的 pending ContentItem（chunk/sentence/mnemonic/syllable）
    - 删除该包的 PackageWord 映射
    - 不删除已生成的内容（非空 content），不删除跨包共享的数据
    """
    from vocab_qc.core.models.enums import MNEMONIC_DIMENSIONS

    # 1. 找出该包关联的所有 word_id
    old_pw_rows = session.query(PackageWord).filter_by(package_id=pkg.id).all()
    old_word_ids = {pw.word_id for pw in old_pw_rows}
    if not old_word_ids:
        return

    # 2. 找出哪些 word 被其他包共享
    shared_word_ids: set[int] = set()
    for wid in old_word_ids:
        other = (
            session.query(PackageWord)
            .filter(PackageWord.word_id == wid, PackageWord.package_id != pkg.id)
            .first()
        )
        if other is not None:
            shared_word_ids.add(wid)

    exclusive_word_ids = old_word_ids - shared_word_ids

    # 3. 找出独占词的所有 meaning_id
    exclusive_meaning_ids: set[int] = set()
    if exclusive_word_ids:
        exclusive_meaning_ids = {
            row[0] for row in
            session.query(Meaning.id).filter(Meaning.word_id.in_(exclusive_word_ids)).all()
        }

    # 4. 收集要删除的 ContentItem IDs，先删关联的 ReviewItem 再删 ContentItem
    ci_ids_to_delete: list[int] = []
    if exclusive_meaning_ids:
        ci_ids_to_delete.extend(
            row[0] for row in session.query(ContentItem.id).filter(
                ContentItem.meaning_id.in_(exclusive_meaning_ids),
                ContentItem.dimension.in_(("chunk", "sentence", *MNEMONIC_DIMENSIONS)),
                ContentItem.qc_status == QcStatus.PENDING.value,
                ContentItem.content == "",
            ).all()
        )

    # 5. 收集独占词的 pending syllable IDs
    if exclusive_word_ids:
        ci_ids_to_delete.extend(
            row[0] for row in session.query(ContentItem.id).filter(
                ContentItem.word_id.in_(exclusive_word_ids),
                ContentItem.dimension == "syllable",
                ContentItem.qc_status == QcStatus.PENDING.value,
                ContentItem.content == "",
            ).all()
        )

    # 6. 先删关联的 ReviewItem，再删 ContentItem（防止孤儿 ReviewItem）
    if ci_ids_to_delete:
        session.query(ReviewItem).filter(
            ReviewItem.content_item_id.in_(ci_ids_to_delete),
        ).delete(synchronize_session=False)
        session.query(ContentItem).filter(
            ContentItem.id.in_(ci_ids_to_delete),
        ).delete(synchronize_session=False)

    # 7. 删除独占词的 RetryCounter
    if exclusive_meaning_ids:
        session.query(RetryCounter).filter(
            RetryCounter.meaning_id.in_(exclusive_meaning_ids),
        ).delete(synchronize_session=False)
    if exclusive_word_ids:
        session.query(RetryCounter).filter(
            RetryCounter.word_id.in_(exclusive_word_ids),
            RetryCounter.meaning_id.is_(None),
        ).delete(synchronize_session=False)

    # 8. 删除 PackageWord 映射
    session.query(PackageWord).filter_by(package_id=pkg.id).delete(synchronize_session=False)
    session.flush()


def _create_content_placeholders(
    session: Session,
    words: list[Word],
    meanings: list[tuple[Word, Meaning]],
) -> None:
    """为导入的数据创建 ContentItem 占位记录。

    chunk / sentence / mnemonic 均按义项生成（防止一词多义混淆）。
    syllable 按单词生成（词级维度，meaning_id=None）。
    """
    from vocab_qc.core.models.enums import MNEMONIC_DIMENSIONS

    per_meaning_dims = ("chunk", "sentence", *MNEMONIC_DIMENSIONS)

    # 批量预查已存在的 ContentItem (word_id, meaning_id, dimension) 三元组
    all_word_ids = list({w.id for w in words})
    existing_triples: set[tuple[int, int | None, str]] = set()

    for i in range(0, len(all_word_ids), 900):
        chunk = all_word_ids[i:i + 900]
        for row in session.query(
            ContentItem.word_id, ContentItem.meaning_id, ContentItem.dimension,
        ).filter(ContentItem.word_id.in_(chunk)).all():
            existing_triples.add((row[0], row[1], row[2]))

    new_items: list[ContentItem] = []

    for word, meaning in meanings:
        for dim in per_meaning_dims:
            if (word.id, meaning.id, dim) not in existing_triples:
                new_items.append(ContentItem(
                    word_id=word.id,
                    meaning_id=meaning.id,
                    dimension=dim,
                    content="",
                    qc_status=QcStatus.PENDING.value,
                ))
                existing_triples.add((word.id, meaning.id, dim))

    # 词级维度: syllable（每个单词一条，meaning_id=None）
    for word in words:
        if (word.id, None, "syllable") not in existing_triples:
            new_items.append(ContentItem(
                word_id=word.id,
                meaning_id=None,
                dimension="syllable",
                content="",
                qc_status=QcStatus.PENDING.value,
            ))
            existing_triples.add((word.id, None, "syllable"))

    # 分批写入 ContentItem，每 5000 条一批，避免大事务锁
    _CONTENT_BATCH_SIZE = 5000
    for i in range(0, len(new_items), _CONTENT_BATCH_SIZE):
        batch = new_items[i:i + _CONTENT_BATCH_SIZE]
        session.add_all(batch)
        session.flush()
        _logger.info(
            "ContentItem 占位写入进度: %d/%d",
            min(i + _CONTENT_BATCH_SIZE, len(new_items)),
            len(new_items),
        )
