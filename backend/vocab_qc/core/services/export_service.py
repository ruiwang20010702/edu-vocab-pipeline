"""导出服务: 仅导出已通过审核的内容."""

import json
import logging
import re
import tempfile
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from vocab_qc.core.logging_config import log_elapsed
from vocab_qc.core.models import (
    ContentItem,
    Meaning,
    Phonetic,
    QcStatus,
    ReviewItem,
    ReviewStatus,
    Source,
    Word,
)
from vocab_qc.core.models.enums import MNEMONIC_DIMENSIONS, QC_TERMINAL_STATUSES

logger = logging.getLogger(__name__)

_MNEMONIC_TYPE_LABELS: dict[str, str] = {
    "mnemonic_root_affix": "词根词缀",
    "mnemonic_word_in_word": "词中词",
    "mnemonic_sound_meaning": "音义联想",
    "mnemonic_exam_app": "考试应用",
}

# Excel 单 sheet 分页阈值：每 N 个义项行（不是单词行）切到下一个 sheet。
# 2 万行 sheet 在普通办公电脑上保持筛选/排序可用；27w 义项约 14 个 sheet。
# 测试可 monkeypatch 改小值验证分页逻辑。
ROWS_PER_SHEET: int = 20000

# 导出层 POS 规范化：仅去尾部点。
# 权威清单（2026-05-26 学科同步）art. 与 det. 并列为合规标签，不再强制 art→det 映射。
_POS_NORMALIZE_MAP: dict[str, str] = {}


def _normalize_pos_for_export(pos: str | None) -> str:
    """导出层规范化 POS：去尾部 '.'，忠于数据库原值。"""
    if not pos:
        return ""
    stripped = pos.rstrip(".")
    return _POS_NORMALIZE_MAP.get(stripped, stripped)


def _format_mnemonic_export(item: "ContentItem") -> dict[str, Any]:
    """助记导出: 解析 JSON content 为结构化数据。"""
    base = {"type": item.dimension}
    try:
        data = json.loads(item.content)
        if isinstance(data, dict):
            base.update(
                {k: data.get(k, "") for k in ("formula", "chant", "extension_words", "script")}
            )
            return base
    except (json.JSONDecodeError, TypeError):
        pass
    base["content"] = item.content
    return base


_MNEMONIC_FIELD_KEYS: tuple[str, ...] = (
    "formula", "chant", "extension_words",
    "exam_sentence", "exam_sentence_translation", "script",
)


def _parse_mnemonic_fields(content: str) -> dict[str, str]:
    """从助记 content 中提取 formula/chant/extension_words/exam_sentence/exam_sentence_translation/script。

    旧数据无 extension_words / exam_sentence / exam_sentence_translation 字段时返回空串。
    """
    empty = dict.fromkeys(_MNEMONIC_FIELD_KEYS, "")
    if not content:
        return empty
    try:
        data = json.loads(content)
        if isinstance(data, dict) and "formula" in data:
            return {k: data.get(k, "") for k in _MNEMONIC_FIELD_KEYS}
    except (json.JSONDecodeError, TypeError):
        pass
    formula = re.search(r"\[核心公式\]\s*([\s\S]*?)(?=\[助记口诀\]|$)", content)
    chant = re.search(r"\[助记口诀\]\s*([\s\S]*?)(?=\[老师话术\]|$)", content)
    script = re.search(r"\[老师话术\]\s*([\s\S]*?)$", content)
    return {
        "formula": (formula.group(1).strip() if formula else ""),
        "chant": (chant.group(1).strip() if chant else ""),
        "extension_words": "",
        "exam_sentence": "",
        "exam_sentence_translation": "",
        "script": (script.group(1).strip() if script else ""),
    }


def _is_meaning_exportable(items: list["ContentItem"]) -> bool:
    """义项可导出条件：所有 ContentItem 都处于终态，且至少有 1 个 approved."""
    if not items:
        return False
    return (
        all(ci.qc_status in QC_TERMINAL_STATUSES for ci in items)
        and any(ci.qc_status == QcStatus.APPROVED.value for ci in items)
    )


def _collect_reviewers_for_meaning(
    session: Session,
    content_item_ids_by_meaning: dict[int, list[int]],
) -> dict[int, str]:
    """按义项聚合最终审核人字符串。

    输入：meaning_id -> [content_item_id, ...]（同一义项下所有维度的 ContentItem.id）
    输出：meaning_id -> "Alice; Bob"（去重排序后 `; ` 拼接，无审核记录返回空串）

    实现要点：
    - 一次性批量查 ReviewItem，避免 N+1
    - 同一 content_item_id 有多条历史 resolved 时取 resolved_at 最新那条（Python 侧去重，兼容 SQLite）
    """
    if not content_item_ids_by_meaning:
        return {}

    all_ci_ids = {ci_id for ids in content_item_ids_by_meaning.values() for ci_id in ids}
    if not all_ci_ids:
        return {}

    # 一次性拉取所有相关的 resolved ReviewItem，按 resolved_at 倒序
    # Python 侧按 content_item_id 去重取首条 reviewer（即最新生效的那位审核员）
    rows = (
        session.query(
            ReviewItem.content_item_id,
            ReviewItem.reviewer,
            ReviewItem.resolved_at,
        )
        .filter(
            ReviewItem.content_item_id.in_(all_ci_ids),
            ReviewItem.status == ReviewStatus.RESOLVED.value,
        )
        .order_by(ReviewItem.resolved_at.desc().nullslast())
        .all()
    )

    reviewer_by_ci: dict[int, str] = {}
    for ci_id, reviewer, _ in rows:
        if ci_id in reviewer_by_ci:
            continue  # 已取过最新那条
        if reviewer:
            reviewer_by_ci[ci_id] = reviewer

    # 按义项聚合：收集该义项下所有维度的 reviewer，去重排序后 `; ` 拼接
    result: dict[int, str] = {}
    for meaning_id, ci_ids in content_item_ids_by_meaning.items():
        names = {reviewer_by_ci[ci_id] for ci_id in ci_ids if ci_id in reviewer_by_ci}
        result[meaning_id] = "; ".join(sorted(names)) if names else ""
    return result


def _iter_approved_batches(session: Session, batch_size: int = 500):
    """分批查询已审核通过的词汇数据（按义项级别过滤：所有维度终态才导出）。

    性能策略：主 ContentItem 查询只 select 导出真正用到的 7 列（id / word_id /
    meaning_id / dimension / qc_status / content / content_cn），避免一并拉
    `generated_with_prompt_*` / `ai_*` / `retry_*` / 时间戳等无关列。Row 对象
    支持属性访问，下游 `.content` / `.dimension` 等调用兼容。
    """
    from collections import defaultdict

    from sqlalchemy import distinct

    # 查出所有有 ContentItem 的 word_id
    all_word_ids = [
        row[0]
        for row in session.query(distinct(ContentItem.word_id)).all()
    ]
    if not all_word_ids:
        return

    for i in range(0, len(all_word_ids), batch_size):
        batch_ids = all_word_ids[i : i + batch_size]

        # 加载该批次所有 ContentItem（不限 status）— 只取 7 列
        stmt = select(
            ContentItem.id,
            ContentItem.word_id,
            ContentItem.meaning_id,
            ContentItem.dimension,
            ContentItem.qc_status,
            ContentItem.content,
            ContentItem.content_cn,
        ).where(ContentItem.word_id.in_(batch_ids))
        all_items = session.execute(stmt).all()

        # 按 (word_id, meaning_id) 分组（all_items 元素为 sqlalchemy.Row，属性访问与 ORM 对象兼容）
        items_by_key: dict[tuple[int, int | None], list[Any]] = defaultdict(list)
        for ci in all_items:
            items_by_key[(ci.word_id, ci.meaning_id)].append(ci)

        # 判定每个 word 的 syllable 是否终态、每个义项是否可导出
        exportable_meaning_ids: set[int] = set()
        word_has_exportable: set[int] = set()
        syllable_ok: dict[int, bool] = {}

        for (word_id, meaning_id), items in items_by_key.items():
            if meaning_id is None:
                # syllable 是 word 级别
                syllable_ok[word_id] = all(
                    ci.qc_status in QC_TERMINAL_STATUSES for ci in items
                )
            else:
                if _is_meaning_exportable(items):
                    exportable_meaning_ids.add(meaning_id)
                    word_has_exportable.add(word_id)

        # 没有 syllable 记录的 word 视为 syllable OK（可能尚未生成）
        exportable_word_ids = [
            wid for wid in batch_ids
            if wid in word_has_exportable and syllable_ok.get(wid, True)
        ]
        if not exportable_word_ids:
            continue

        words = {w.id: w for w in session.query(Word).filter(Word.id.in_(exportable_word_ids)).all()}
        phonetics = {}
        for p in session.query(Phonetic).filter(Phonetic.word_id.in_(exportable_word_ids)).all():
            phonetics[p.word_id] = p

        all_meanings = session.query(Meaning).filter(Meaning.word_id.in_(exportable_word_ids)).all()
        meanings_by_word: dict[int, list[Meaning]] = defaultdict(list)
        meaning_ids = []
        for m in all_meanings:
            if m.id in exportable_meaning_ids:
                meanings_by_word[m.word_id].append(m)
                meaning_ids.append(m.id)

        sources_by_meaning: dict[int, list[Source]] = defaultdict(list)
        if meaning_ids:
            for s in session.query(Source).filter(Source.meaning_id.in_(meaning_ids)).all():
                sources_by_meaning[s.meaning_id].append(s)

        # 收集每个可导出义项下所有维度的 ContentItem.id，用于聚合审核人
        # syllable（meaning_id=None）属于词级，不参与义项聚合
        ci_ids_by_meaning: dict[int, list[int]] = defaultdict(list)
        for ci in all_items:
            if ci.meaning_id and ci.meaning_id in exportable_meaning_ids:
                ci_ids_by_meaning[ci.meaning_id].append(ci.id)
        reviewers_by_meaning = _collect_reviewers_for_meaning(session, dict(ci_ids_by_meaning))

        # 只取 approved 的 ContentItem 填充内容
        approved_items = [ci for ci in all_items if ci.qc_status == QcStatus.APPROVED.value]

        content_index: dict[tuple[int, int | None, str], Any] = {}
        mnemonics_by_meaning: dict[int, list[Any]] = defaultdict(list)
        for ci in approved_items:
            if ci.word_id not in words:
                continue
            if ci.dimension in MNEMONIC_DIMENSIONS and ci.meaning_id:
                if ci.meaning_id in exportable_meaning_ids:
                    mnemonics_by_meaning[ci.meaning_id].append(ci)
            else:
                content_index[(ci.word_id, ci.meaning_id, ci.dimension)] = ci

        for word_id in exportable_word_ids:
            word = words.get(word_id)
            if not word:
                continue

            export_meanings = meanings_by_word.get(word_id, [])
            if not export_meanings:
                continue

            phonetic = phonetics.get(word_id)
            syllable_item = content_index.get((word_id, None, "syllable"))
            syllables = (
                syllable_item.content if syllable_item
                else (phonetic.syllables if phonetic else "")
            )
            result: dict[str, Any] = {
                "id": word.id,
                "word": word.word,
                "syllables": syllables,
                "ipa_uk": phonetic.ipa_uk if phonetic else "",
                "ipa_us": phonetic.ipa_us if phonetic else "",
                "audio_url_uk": phonetic.audio_url_uk if phonetic else "",
                "audio_url_us": phonetic.audio_url_us if phonetic else "",
                "meanings": [],
            }

            for meaning in export_meanings:
                sources = sources_by_meaning.get(meaning.id, [])
                chunk = content_index.get((word_id, meaning.id, "chunk"))
                sentence = content_index.get((word_id, meaning.id, "sentence"))

                first_source = sources[0] if sources else None
                meaning_data = {
                    "pos": _normalize_pos_for_export(meaning.pos),
                    "def": meaning.definition,
                    "sources": [s.source_name for s in sources],
                    "textbook_id": first_source.textbook_id if first_source else None,
                    "word_book_id": first_source.word_book_id if first_source else None,
                    "unit_id": first_source.unit_id if first_source else None,
                    "chunk": chunk.content if chunk else None,
                    "chunk_cn": chunk.content_cn if chunk else None,
                    "sentence": sentence.content if sentence else None,
                    "sentence_cn": sentence.content_cn if sentence else None,
                    "mnemonics": [
                        {"type": m.dimension, "content": m.content}
                        for m in mnemonics_by_meaning.get(meaning.id, [])
                    ],
                    "reviewer": reviewers_by_meaning.get(meaning.id, ""),
                }
                result["meanings"].append(meaning_data)

            yield result


class ExportService:
    """导出服务: 门禁 + 格式化输出."""

    def export_word(self, session: Session, word_id: int) -> dict[str, Any] | None:
        """导出单个词的完整数据（按义项过滤：所有维度终态才导出）."""
        from collections import defaultdict

        logger.info("导出单词开始 word_id=%d", word_id)

        word = session.query(Word).filter_by(id=word_id).first()
        if not word:
            logger.warning("导出单词未找到 word_id=%d", word_id)
            return None

        phonetic = session.query(Phonetic).filter_by(word_id=word.id).first()
        meanings = session.query(Meaning).filter_by(word_id=word.id).all()

        # 加载该词所有 ContentItem（不限 status）
        all_items = (
            session.query(ContentItem)
            .filter(ContentItem.word_id == word.id)
            .all()
        )

        # 按 meaning_id 分组，判定哪些义项可导出
        items_by_meaning: dict[int | None, list[ContentItem]] = defaultdict(list)
        for ci in all_items:
            items_by_meaning[ci.meaning_id].append(ci)

        # syllable 终态检查
        syllable_items = items_by_meaning.get(None, [])
        if syllable_items and not all(ci.qc_status in QC_TERMINAL_STATUSES for ci in syllable_items):
            logger.info("导出单词跳过 word_id=%d syllable 未终态", word_id)
            return None

        # 筛选可导出义项
        exportable_meaning_ids = {
            mid for mid, items in items_by_meaning.items()
            if mid is not None and _is_meaning_exportable(items)
        }

        if not exportable_meaning_ids:
            logger.info("导出单词跳过 word_id=%d 无可导出义项", word_id)
            return None

        # 只取 approved 的 ContentItem 填充内容
        approved_items = [ci for ci in all_items if ci.qc_status == QcStatus.APPROVED.value]

        content_index: dict[tuple[int | None, str], ContentItem] = {}
        mnemonics_by_meaning: dict[int, list[ContentItem]] = defaultdict(list)
        for ci in approved_items:
            if ci.dimension in MNEMONIC_DIMENSIONS and ci.meaning_id:
                if ci.meaning_id in exportable_meaning_ids:
                    mnemonics_by_meaning[ci.meaning_id].append(ci)
            else:
                content_index[(ci.meaning_id, ci.dimension)] = ci

        meaning_ids = [m.id for m in meanings if m.id in exportable_meaning_ids]
        sources_by_meaning: dict[int, list[Source]] = defaultdict(list)
        if meaning_ids:
            for s in session.query(Source).filter(Source.meaning_id.in_(meaning_ids)).all():
                sources_by_meaning[s.meaning_id].append(s)

        syllable_item = content_index.get((None, "syllable"))
        syllables = (
            syllable_item.content if syllable_item
            else (phonetic.syllables if phonetic else "")
        )

        result: dict[str, Any] = {
            "id": word.id,
            "word": word.word,
            "syllables": syllables,
            "ipa_uk": phonetic.ipa_uk if phonetic else "",
            "ipa_us": phonetic.ipa_us if phonetic else "",
            "audio_url_uk": phonetic.audio_url_uk if phonetic else "",
            "audio_url_us": phonetic.audio_url_us if phonetic else "",
            "meanings": [],
        }

        for meaning in meanings:
            if meaning.id not in exportable_meaning_ids:
                continue

            chunk = content_index.get((meaning.id, "chunk"))
            sentence = content_index.get((meaning.id, "sentence"))

            sources = sources_by_meaning.get(meaning.id, [])
            first_source = sources[0] if sources else None
            meaning_data = {
                "pos": _normalize_pos_for_export(meaning.pos),
                "def": meaning.definition,
                "sources": [s.source_name for s in sources],
                "textbook_id": first_source.textbook_id if first_source else None,
                "word_book_id": first_source.word_book_id if first_source else None,
                "unit_id": first_source.unit_id if first_source else None,
                "chunk": chunk.content if chunk else None,
                "chunk_cn": chunk.content_cn if chunk else None,
                "sentence": sentence.content if sentence else None,
                "sentence_cn": sentence.content_cn if sentence else None,
                "mnemonics": [_format_mnemonic_export(m) for m in mnemonics_by_meaning.get(meaning.id, [])],
            }
            result["meanings"].append(meaning_data)

        return result

    def export_all_approved(self, session: Session) -> list[dict[str, Any]]:
        """导出所有有 approved 内容的词（复用分批生成器，避免全量加载）."""
        logger.info("导出全部已审核词汇开始")
        with log_elapsed(logger, "导出全部已审核词汇"):
            result = list(_iter_approved_batches(session))
        logger.info("导出全部已审核词汇完成 count=%d", len(result))
        return result

    def export_to_json(self, session: Session, filepath: str) -> int:
        """导出到 JSON 文件（分批写入，避免全量加载到内存）."""
        logger.info("导出 JSON 开始 filepath=%s", filepath)
        count = 0
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("[\n")
            for item in _iter_approved_batches(session):
                if count > 0:
                    f.write(",\n")
                json.dump(item, f, ensure_ascii=False, indent=2)
                count += 1
            f.write("\n]")
        logger.info("导出 JSON 完成 filepath=%s count=%d", filepath, count)
        return count

    def export_to_excel(
        self,
        session: Session,
        output_path: Path | None = None,
    ) -> Path:
        """导出已通过词汇数据为 Excel，每个义项一行，4 种助记按维度展列。

        P-M1: 分批查询 + openpyxl write_only 流式写入，避免在内存里构建整本工作簿对象树。
        P-M2: 写入到临时文件而非 BytesIO，避免在内存中完整持有压缩后的 xlsx 二进制。
        P-M3: 每 ROWS_PER_SHEET (默认 2 万) 个义项行切到下一个 sheet，单 sheet 保持
              筛选/排序可用；27 万义项约 14 个 sheet。

        Args:
            output_path: 落盘路径。None 时自动创建带 .xlsx 后缀的 NamedTemporaryFile，
                调用方负责清理（API router 用 BackgroundTask）。
        Returns:
            写入完成的 xlsx 文件 Path。
        """
        logger.info("导出 Excel 开始")
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # 区分 service 自创的 tempfile 与调用方传入的 output_path：
        # 仅在自创路径上做异常 unlink（不擅自删除调用方文件）。
        _owned_tempfile = False
        if output_path is None:
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            output_path = Path(tmp.name)
            tmp.close()
            _owned_tempfile = True

        try:
            return self._write_xlsx(session, output_path)
        except Exception:
            # 失败时清理自创临时文件，避免 /tmp 累积；调用方文件保留交由其处理
            if _owned_tempfile:
                try:
                    output_path.unlink(missing_ok=True)
                except OSError:
                    logger.exception("清理失败的临时 xlsx 文件出错 path=%s", output_path)
            raise

    def _write_xlsx(self, session: Session, output_path: Path) -> Path:
        """构建 workbook 并落盘到 output_path（被 export_to_excel 包 try/except 调用）。"""
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # 助记类型与各自要导出的列：通用 3 列；词根词缀额外含"同根词"为 4 列；考试应用含"例句"为 5 列
        basic_mn_cols: list[tuple[str, str]] = [
            ("formula", "公式"), ("chant", "口诀"), ("script", "话术"),
        ]
        root_affix_mn_cols: list[tuple[str, str]] = [
            ("formula", "公式"), ("chant", "口诀"),
            ("extension_words", "同根词"), ("script", "话术"),
        ]
        exam_mn_cols: list[tuple[str, str]] = [
            ("formula", "公式"), ("chant", "口诀"),
            ("exam_sentence", "例句"), ("exam_sentence_translation", "例句释义"),
            ("script", "话术"),
        ]
        mnemonic_types: list[tuple[str, str, list[tuple[str, str]]]] = [
            ("mnemonic_root_affix", "词根词缀", root_affix_mn_cols),
            ("mnemonic_word_in_word", "词中词", basic_mn_cols),
            ("mnemonic_sound_meaning", "谐音联想", basic_mn_cols),
            ("mnemonic_exam_app", "考试应用", exam_mn_cols),
        ]

        base_headers = [
            "单词", "英式音标", "美式音标", "英式音频URL", "美式音频URL",
            "音节", "词性", "释义", "教材来源",
            "教材ID", "词书ID", "单元ID",
            "语块", "语块翻译", "例句", "例句翻译",
        ]
        mn_headers: list[str] = []
        for _, label, cols in mnemonic_types:
            for _, suffix in cols:
                mn_headers.append(f"{label}·{suffix}")
        headers = base_headers + mn_headers + ["审核人"]

        # write_only 工作簿不能用 wb.active；每个 sheet 通过 create_sheet 创建
        wb = Workbook(write_only=True)

        # 列宽（每个 sheet 独立配置）
        base_widths = [15, 22, 22, 40, 40, 18, 8, 30, 24, 12, 12, 12, 28, 28, 42, 42]
        basic_mn_widths: list[int] = [22, 22, 30]
        root_affix_mn_widths: list[int] = [22, 22, 28, 30]
        exam_mn_widths: list[int] = [22, 22, 35, 30, 30]
        mn_widths: list[int] = []
        for mn_key, _, _ in mnemonic_types:
            if mn_key == "mnemonic_exam_app":
                mn_widths += exam_mn_widths
            elif mn_key == "mnemonic_root_affix":
                mn_widths += root_affix_mn_widths
            else:
                mn_widths += basic_mn_widths
        col_widths = base_widths + mn_widths + [20]  # 末列：审核人

        # 样式：所有 sheet 共享同一组样式实例（openpyxl 内部按 hash 去重）
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        wrap_align = Alignment(wrap_text=True, vertical="top")

        def _new_sheet(idx: int):
            """创建一个新 sheet，写入列宽、freeze panes、表头样式。"""
            ws_ = wb.create_sheet(f"词表导出_{idx:02d}")
            for i, w in enumerate(col_widths, 1):
                ws_.column_dimensions[get_column_letter(i)].width = w
            ws_.freeze_panes = "A2"
            header_cells = []
            for h in headers:
                c = WriteOnlyCell(ws_, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = wrap_align
                header_cells.append(c)
            ws_.append(header_cells)
            return ws_

        def _wrap_row(ws_, values: list[Any]) -> list[Any]:
            """把一行的值用 WriteOnlyCell + 共享 wrap_align 包一层，保留自动换行。"""
            cells = []
            for v in values:
                c = WriteOnlyCell(ws_, value=v)
                c.alignment = wrap_align
                cells.append(c)
            return cells

        sheet_idx = 1
        ws = _new_sheet(sheet_idx)
        rows_in_sheet = 0
        total_rows = 0

        def _maybe_rotate_sheet():
            nonlocal ws, sheet_idx, rows_in_sheet
            if rows_in_sheet >= ROWS_PER_SHEET:
                sheet_idx += 1
                ws = _new_sheet(sheet_idx)
                rows_in_sheet = 0

        for word_data in _iter_approved_batches(session):
            word = word_data["word"]
            ipa_uk = word_data.get("ipa_uk", "")
            ipa_us = word_data.get("ipa_us", "")
            audio_uk = word_data.get("audio_url_uk", "") or ""
            audio_us = word_data.get("audio_url_us", "") or ""
            syllables = word_data.get("syllables", "")
            meanings = word_data.get("meanings", [])

            if not meanings:
                _maybe_rotate_sheet()
                row_values: list[Any] = [word, ipa_uk, ipa_us, audio_uk, audio_us, syllables]
                row_values += [""] * (len(headers) - len(row_values))
                ws.append(_wrap_row(ws, row_values))
                rows_in_sheet += 1
                total_rows += 1
                continue

            for m in meanings:
                _maybe_rotate_sheet()
                sources = m.get("sources", [])
                row_values = [
                    word, ipa_uk, ipa_us, audio_uk, audio_us, syllables,
                    m.get("pos", ""),
                    m.get("def", ""),
                    "; ".join(sources) if sources else "",
                    m.get("textbook_id") or "",
                    m.get("word_book_id") or "",
                    m.get("unit_id") or "",
                    m.get("chunk") or "",
                    m.get("chunk_cn") or "",
                    m.get("sentence") or "",
                    m.get("sentence_cn") or "",
                ]

                # 助记：按 type 建索引
                mn_by_type: dict[str, dict[str, str]] = {}
                for mn in m.get("mnemonics", []):
                    mn_by_type[mn["type"]] = _parse_mnemonic_fields(mn.get("content", ""))

                # 各类型按其 columns 追加，缺失（LLM 判定 valid:false / 该维度 rejected）留空
                for mn_key, _, cols in mnemonic_types:
                    fields = mn_by_type.get(mn_key)
                    for field_name, _ in cols:
                        row_values.append(fields[field_name] if fields else "")

                # 审核人列（最后一列）
                row_values.append(m.get("reviewer", ""))

                ws.append(_wrap_row(ws, row_values))
                rows_in_sheet += 1
                total_rows += 1

        wb.save(str(output_path))
        logger.info("导出 Excel 完成 rows=%d sheets=%d path=%s", total_rows, sheet_idx, output_path)
        return output_path

    def get_export_readiness(self, session: Session) -> dict:
        """检查导出就绪状态."""
        total = session.query(ContentItem).count()
        approved = session.query(ContentItem).filter_by(qc_status=QcStatus.APPROVED.value).count()
        pending = session.query(ContentItem).filter_by(qc_status=QcStatus.PENDING.value).count()

        return {
            "total_items": total,
            "approved": approved,
            "pending": pending,
            "not_approved": total - approved,
            "ready_rate": round(approved / total * 100, 1) if total > 0 else 0,
        }
