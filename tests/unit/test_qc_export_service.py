"""QcService 与 ExportService 单元测试."""

from datetime import UTC, datetime, timedelta

import pytest
from sqlalchemy.orm import Session
from vocab_qc.core.models import (
    ContentItem,
    Meaning,
    QcStatus,
    ReviewItem,
    ReviewReason,
    ReviewStatus,
    Word,
)
from vocab_qc.core.services.export_service import ExportService
from vocab_qc.core.services.qc_service import QcService

# ---- helpers ----


def _make_word(session: Session, word_text: str) -> Word:
    word = Word(word=word_text)
    session.add(word)
    session.flush()
    return word


def _make_meaning(session: Session, word: Word, pos: str, definition: str) -> Meaning:
    meaning = Meaning(word_id=word.id, pos=pos, definition=definition)
    session.add(meaning)
    session.flush()
    return meaning


def _make_content(
    session: Session,
    word: Word,
    dimension: str,
    content: str,
    *,
    meaning: Meaning | None = None,
    qc_status: str = QcStatus.PENDING.value,
    content_cn: str | None = None,
) -> ContentItem:
    item = ContentItem(
        word_id=word.id,
        meaning_id=meaning.id if meaning else None,
        dimension=dimension,
        content=content,
        content_cn=content_cn,
        qc_status=qc_status,
    )
    session.add(item)
    session.flush()
    return item


# ---- QcService 测试 ----


class TestQcServiceInvalidScope:
    """run_layer1 使用不匹配任何项的 dimension 参数时的行为."""

    def test_dimension_matches_no_items(self, db_session: Session):
        """当 dimension 筛选匹配不到任何 ContentItem 时，返回零结果."""
        word = _make_word(db_session, "apple")
        meaning = _make_meaning(db_session, word, "n.", "苹果")
        _make_content(db_session, word, "chunk", "a red apple", meaning=meaning)

        svc = QcService()
        result = svc.run_layer1(db_session, dimension="sentence")

        assert result == {"run_id": None, "total": 0, "passed": 0, "failed": 0}

    def test_nonexistent_dimension(self, db_session: Session):
        """当 dimension 为完全不存在的值时，返回零结果."""
        word = _make_word(db_session, "banana")
        meaning = _make_meaning(db_session, word, "n.", "香蕉")
        _make_content(db_session, word, "chunk", "a banana", meaning=meaning)

        svc = QcService()
        result = svc.run_layer1(db_session, dimension="nonexistent_dim")

        assert result["run_id"] is None
        assert result["total"] == 0


class TestQcServiceTerminalStatusFiltering:
    """run_layer1 跳过 APPROVED / REJECTED 终态项."""

    def test_approved_items_are_skipped(self, db_session: Session):
        """APPROVED 状态的内容项不应被 run_layer1 处理."""
        word = _make_word(db_session, "cat")
        meaning = _make_meaning(db_session, word, "n.", "猫")
        _make_content(
            db_session, word, "chunk", "a lovely cat",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )

        svc = QcService()
        result = svc.run_layer1(db_session)

        assert result["total"] == 0
        assert result["run_id"] is None

    def test_rejected_items_are_skipped(self, db_session: Session):
        """REJECTED 状态的内容项不应被 run_layer1 处理."""
        word = _make_word(db_session, "dog")
        meaning = _make_meaning(db_session, word, "n.", "狗")
        _make_content(
            db_session, word, "chunk", "a big dog",
            meaning=meaning, qc_status=QcStatus.REJECTED.value,
        )

        svc = QcService()
        result = svc.run_layer1(db_session)

        assert result["total"] == 0
        assert result["run_id"] is None

    def test_mix_terminal_and_pending(self, db_session: Session):
        """混合终态与待处理项，仅待处理项被纳入质检."""
        word = _make_word(db_session, "run")
        meaning = _make_meaning(db_session, word, "v.", "跑")

        _make_content(
            db_session, word, "chunk", "run fast",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        _make_content(
            db_session, word, "chunk", "run away",
            meaning=meaning, qc_status=QcStatus.REJECTED.value,
        )
        _make_content(
            db_session, word, "chunk", "run a race",
            meaning=meaning, qc_status=QcStatus.PENDING.value,
        )

        svc = QcService()
        result = svc.run_layer1(db_session)

        # 只有 pending 的那一项被处理
        assert result["total"] == 1


# ---- ExportService 测试 ----


class TestExportServiceAllApproved:
    """export_all_approved 按义项级别过滤：所有维度终态才导出."""

    def test_meaning_with_pending_item_not_exported(self, db_session: Session):
        """义项有 pending 维度时，整个义项不导出."""
        word = _make_word(db_session, "book")
        meaning = _make_meaning(db_session, word, "n.", "书")

        _make_content(
            db_session, word, "chunk", "read a book",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        # pending sentence → 义项未全部终态 → 不导出
        _make_content(
            db_session, word, "sentence", "I read a book every day.",
            meaning=meaning, qc_status=QcStatus.PENDING.value,
            content_cn="我每天读一本书。",
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)
        assert results == []

    def test_all_terminal_meaning_exported(self, db_session: Session):
        """义项所有维度均终态（approved + rejected），可导出."""
        word = _make_word(db_session, "book")
        meaning = _make_meaning(db_session, word, "n.", "书")

        _make_content(
            db_session, word, "chunk", "read a book",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        # rejected = 不适用，属于终态
        _make_content(
            db_session, word, "mnemonic_word_in_word", "{}",
            meaning=meaning, qc_status=QcStatus.REJECTED.value,
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)
        assert len(results) == 1
        assert results[0]["word"] == "book"
        assert results[0]["meanings"][0]["chunk"] == "read a book"

    def test_all_rejected_meaning_not_exported(self, db_session: Session):
        """义项所有维度都是 rejected（全部不适用），不导出（无有效内容）."""
        word = _make_word(db_session, "a")
        meaning = _make_meaning(db_session, word, "art.", "一个")

        _make_content(
            db_session, word, "mnemonic_root_affix", "{}",
            meaning=meaning, qc_status=QcStatus.REJECTED.value,
        )
        _make_content(
            db_session, word, "mnemonic_word_in_word", "{}",
            meaning=meaning, qc_status=QcStatus.REJECTED.value,
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)
        assert results == []

    def test_no_approved_items(self, db_session: Session):
        """没有任何 approved 项时返回空列表."""
        word = _make_word(db_session, "pen")
        meaning = _make_meaning(db_session, word, "n.", "笔")
        _make_content(
            db_session, word, "chunk", "a pen",
            meaning=meaning, qc_status=QcStatus.PENDING.value,
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)
        assert results == []

    def test_multiple_words_partial_approval(self, db_session: Session):
        """多词场景：仅义项全部终态的词才出现在导出列表中."""
        # sun: 全部终态 → 可导出
        word_a = _make_word(db_session, "sun")
        meaning_a = _make_meaning(db_session, word_a, "n.", "太阳")
        _make_content(
            db_session, word_a, "chunk", "the sun rises",
            meaning=meaning_a, qc_status=QcStatus.APPROVED.value,
        )

        # moon: pending → 不可导出
        word_b = _make_word(db_session, "moon")
        meaning_b = _make_meaning(db_session, word_b, "n.", "月亮")
        _make_content(
            db_session, word_b, "chunk", "the moon",
            meaning=meaning_b, qc_status=QcStatus.PENDING.value,
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)

        exported_words = [r["word"] for r in results]
        assert "sun" in exported_words
        assert "moon" not in exported_words

    def test_multi_meaning_partial_export(self, db_session: Session):
        """一个词有两个义项，只导出全部终态的义项."""
        word = _make_word(db_session, "run")
        m1 = _make_meaning(db_session, word, "v.", "跑")
        m2 = _make_meaning(db_session, word, "v.", "运行")

        # m1: 全部 approved → 可导出
        _make_content(
            db_session, word, "chunk", "run fast",
            meaning=m1, qc_status=QcStatus.APPROVED.value,
        )
        # m2: 有 pending → 不可导出
        _make_content(
            db_session, word, "chunk", "run a program",
            meaning=m2, qc_status=QcStatus.APPROVED.value,
        )
        _make_content(
            db_session, word, "sentence", "I run the program.",
            meaning=m2, qc_status=QcStatus.PENDING.value,
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)

        assert len(results) == 1
        assert len(results[0]["meanings"]) == 1
        assert results[0]["meanings"][0]["def"] == "跑"

    def test_syllable_not_terminal_blocks_word(self, db_session: Session):
        """syllable 未终态时，整个词不导出."""
        word = _make_word(db_session, "cat")
        meaning = _make_meaning(db_session, word, "n.", "猫")

        # 义项全部 approved
        _make_content(
            db_session, word, "chunk", "a lovely cat",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        # syllable pending → 阻塞导出
        _make_content(
            db_session, word, "syllable", "cat",
            meaning=None, qc_status=QcStatus.PENDING.value,
        )

        svc = ExportService()
        results = svc.export_all_approved(db_session)
        assert results == []

    def test_export_readiness_counts(self, db_session: Session):
        """get_export_readiness 正确统计各状态数量."""
        word = _make_word(db_session, "star")
        meaning = _make_meaning(db_session, word, "n.", "星星")
        _make_content(
            db_session, word, "chunk", "a star",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        _make_content(
            db_session, word, "sentence", "I see a star.",
            meaning=meaning, qc_status=QcStatus.PENDING.value,
            content_cn="我看到一颗星星。",
        )
        _make_content(
            db_session, word, "chunk", "bad star",
            meaning=meaning, qc_status=QcStatus.REJECTED.value,
        )

        svc = ExportService()
        readiness = svc.get_export_readiness(db_session)

        assert readiness["total_items"] == 3
        assert readiness["approved"] == 1
        assert readiness["pending"] == 1
        assert readiness["not_approved"] == 2
        assert readiness["ready_rate"] == pytest.approx(33.3, abs=0.1)


class TestExportPosNormalization:
    """导出层 POS 规范化端到端：旧带点 / art. 在导出时变换；新格式保持不变."""

    def _seed_exportable_word(
        self,
        session: Session,
        word_text: str,
        pos: str,
        definition: str,
    ) -> Word:
        """造一个可导出的最小词（chunk approved 即可触发义项 exportable）."""
        word = _make_word(session, word_text)
        meaning = _make_meaning(session, word, pos, definition)
        _make_content(
            session, word, "chunk", f"some {word_text}",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        return word

    def test_legacy_dotted_pos_stripped_in_export_all(self, db_session: Session):
        """export_all_approved: 'n.' / 'adj.' 等带点格式去点后输出."""
        self._seed_exportable_word(db_session, "book", "n.", "书")
        self._seed_exportable_word(db_session, "kind", "adj.", "友好的")

        results = ExportService().export_all_approved(db_session)
        pos_values = {r["word"]: r["meanings"][0]["pos"] for r in results}

        assert pos_values["book"] == "n"
        assert pos_values["kind"] == "adj"

    def test_art_preserved_in_export_all(self, db_session: Session):
        """export_all_approved: 'art.' 去点后保持为 'art'（不再映射到 det）。

        2026-05-26 权威清单 art. 与 det. 并存，删除历史 art→det 映射。
        """
        self._seed_exportable_word(db_session, "the", "art.", "这个")

        results = ExportService().export_all_approved(db_session)
        assert len(results) == 1
        assert results[0]["meanings"][0]["pos"] == "art"

    def test_new_format_passthrough_in_export_all(self, db_session: Session):
        """export_all_approved: 新格式 'n phr' / 'a phr' / 'n' 保持不变."""
        self._seed_exportable_word(db_session, "look forward to", "n phr", "期待")
        self._seed_exportable_word(db_session, "lovely to look at", "a phr", "悦目的")
        self._seed_exportable_word(db_session, "cat", "n", "猫")

        results = ExportService().export_all_approved(db_session)
        pos_values = {r["word"]: r["meanings"][0]["pos"] for r in results}

        assert pos_values["look forward to"] == "n phr"
        assert pos_values["lovely to look at"] == "a phr"
        assert pos_values["cat"] == "n"

    def test_export_word_path_normalizes_pos(self, db_session: Session):
        """export_word 单词路径独立规范化：去点保持原值。"""
        word = self._seed_exportable_word(db_session, "an", "art.", "一个")

        result = ExportService().export_word(db_session, word.id)
        assert result is not None
        assert result["meanings"][0]["pos"] == "art"


def _make_review_item(
    session: Session,
    content_item: ContentItem,
    *,
    reviewer: str,
    resolved_at: datetime | None = None,
    status: str = ReviewStatus.RESOLVED.value,
) -> ReviewItem:
    """构造一个 resolved ReviewItem 用于审核人列测试。"""
    ri = ReviewItem(
        content_item_id=content_item.id,
        word_id=content_item.word_id,
        meaning_id=content_item.meaning_id,
        dimension=content_item.dimension,
        reason=ReviewReason.LAYER1_FAILED.value,
        status=status,
        reviewer=reviewer,
        resolved_at=resolved_at or datetime.now(UTC),
    )
    session.add(ri)
    session.flush()
    return ri


class TestExportReviewerColumn:
    """export_all_approved 注入审核人字段，按义项聚合去重排序。"""

    def test_reviewer_single_dimension_single_reviewer(self, db_session: Session):
        """单义项单维度单审核员 → reviewer 字段 = 单个名字。"""
        word = _make_word(db_session, "book")
        meaning = _make_meaning(db_session, word, "n", "书")
        chunk = _make_content(
            db_session, word, "chunk", "read a book",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        _make_review_item(db_session, chunk, reviewer="alice")

        results = ExportService().export_all_approved(db_session)
        assert len(results) == 1
        assert results[0]["meanings"][0]["reviewer"] == "alice"

    def test_reviewer_multi_dim_same_user_dedup(self, db_session: Session):
        """同义项多维度同审核员 → 去重为单个名字（典型场景：词维度派发）。"""
        word = _make_word(db_session, "run")
        meaning = _make_meaning(db_session, word, "v", "跑")

        chunk = _make_content(
            db_session, word, "chunk", "run fast",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        sentence = _make_content(
            db_session, word, "sentence", "I run fast.",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        _make_review_item(db_session, chunk, reviewer="bob")
        _make_review_item(db_session, sentence, reviewer="bob")

        results = ExportService().export_all_approved(db_session)
        assert len(results) == 1
        assert results[0]["meanings"][0]["reviewer"] == "bob"

    def test_reviewer_multi_dim_diff_users_joined(self, db_session: Session):
        """同义项多维度不同审核员（罕见跨批次场景）→ 排序后 `; ` 拼接。"""
        word = _make_word(db_session, "fly")
        meaning = _make_meaning(db_session, word, "v", "飞")

        chunk = _make_content(
            db_session, word, "chunk", "fly high",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        sentence = _make_content(
            db_session, word, "sentence", "Birds fly high.",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        _make_review_item(db_session, chunk, reviewer="zoe")
        _make_review_item(db_session, sentence, reviewer="alice")

        results = ExportService().export_all_approved(db_session)
        assert len(results) == 1
        # 排序后应为 alice; zoe
        assert results[0]["meanings"][0]["reviewer"] == "alice; zoe"

    def test_reviewer_isolated_per_meaning(self, db_session: Session):
        """不同义项的 reviewer 互不污染（验证按 meaning_id 隔离聚合）。

        说明：同一 ContentItem 历史多条 resolved 取最新那条，靠 `_collect_reviewers_for_meaning`
        内的 `ORDER BY resolved_at DESC` + Python 侧首条去重保证；该路径在 SQLite 测试环境下
        受 partial unique index 退化限制无法构造，由 PostgreSQL 集成验证。
        """
        word = _make_word(db_session, "open")
        m1 = _make_meaning(db_session, word, "v", "打开")
        m2 = _make_meaning(db_session, word, "v", "开始")

        chunk1 = _make_content(
            db_session, word, "chunk", "open the door",
            meaning=m1, qc_status=QcStatus.APPROVED.value,
        )
        chunk2 = _make_content(
            db_session, word, "chunk", "open a new chapter",
            meaning=m2, qc_status=QcStatus.APPROVED.value,
        )

        old_time = datetime.now(UTC) - timedelta(days=1)
        new_time = datetime.now(UTC)
        _make_review_item(db_session, chunk1, reviewer="alice", resolved_at=old_time)
        _make_review_item(db_session, chunk2, reviewer="bob", resolved_at=new_time)

        results = ExportService().export_all_approved(db_session)
        assert len(results) == 1
        meanings_by_def = {m["def"]: m["reviewer"] for m in results[0]["meanings"]}
        assert meanings_by_def["打开"] == "alice"
        assert meanings_by_def["开始"] == "bob"

    def test_reviewer_missing_review_returns_empty(self, db_session: Session):
        """ContentItem 无 ReviewItem 时 reviewer 字段为空字符串。"""
        word = _make_word(db_session, "pen")
        meaning = _make_meaning(db_session, word, "n", "笔")
        _make_content(
            db_session, word, "chunk", "a pen",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        # 故意不创建 ReviewItem

        results = ExportService().export_all_approved(db_session)
        assert len(results) == 1
        assert results[0]["meanings"][0]["reviewer"] == ""

    def test_excel_export_reviewer_header_and_value(self, db_session: Session):
        """端到端验证 Excel 文件最后一列表头为「审核人」且数据正确。"""
        from openpyxl import load_workbook

        word = _make_word(db_session, "cat")
        meaning = _make_meaning(db_session, word, "n", "猫")
        chunk = _make_content(
            db_session, word, "chunk", "a lovely cat",
            meaning=meaning, qc_status=QcStatus.APPROVED.value,
        )
        _make_review_item(db_session, chunk, reviewer="wangrui003")

        buf = ExportService().export_to_excel(db_session)
        wb = load_workbook(buf)
        ws = wb.active

        # 表头最后一列
        last_col_idx = ws.max_column
        assert ws.cell(row=1, column=last_col_idx).value == "审核人"

        # 数据行最后一列：找到 cat 行
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == "cat":
                assert ws.cell(row=row_idx, column=last_col_idx).value == "wangrui003"
                break
        else:
            raise AssertionError("未找到 cat 数据行")


class TestExcelPagination:
    """多 sheet 分页：每 ROWS_PER_SHEET 个义项切到下一个 sheet（27w 义项约 14 个 sheet）。"""

    def test_paginates_when_rows_exceed_threshold(
        self, db_session: Session, monkeypatch: pytest.MonkeyPatch,
    ):
        """ROWS_PER_SHEET=2 + 3 个义项 → 2 个 sheet（2+1 分布），每 sheet 都有完整表头。"""
        from openpyxl import load_workbook
        from vocab_qc.core.services import export_service as export_mod

        monkeypatch.setattr(export_mod, "ROWS_PER_SHEET", 2)

        for w in ("alpha", "beta", "gamma"):
            word = _make_word(db_session, w)
            meaning = _make_meaning(db_session, word, "n.", f"def-{w}")
            _make_content(
                db_session, word, "chunk", f"chunk-{w}",
                meaning=meaning, qc_status=QcStatus.APPROVED.value,
            )

        path = ExportService().export_to_excel(db_session)
        wb = load_workbook(path)

        assert wb.sheetnames == ["词表导出_01", "词表导出_02"], (
            f"期望 2 个 sheet 命名 词表导出_01/02，实得 {wb.sheetnames}"
        )

        # sheet 01：表头 + 2 行数据
        ws1 = wb["词表导出_01"]
        assert ws1.cell(row=1, column=1).value == "单词"
        assert ws1.cell(row=1, column=ws1.max_column).value == "审核人"
        assert ws1.max_row == 3  # 1 表头 + 2 数据

        # sheet 02：表头 + 1 行数据
        ws2 = wb["词表导出_02"]
        assert ws2.cell(row=1, column=1).value == "单词"
        assert ws2.cell(row=1, column=ws2.max_column).value == "审核人"
        assert ws2.max_row == 2  # 1 表头 + 1 数据

        # 3 个单词分布在两个 sheet 里（顺序不强求，只要齐全）
        words_seen = set()
        for ws in (ws1, ws2):
            for row_idx in range(2, ws.max_row + 1):
                w = ws.cell(row=row_idx, column=1).value
                if w:
                    words_seen.add(w)
        assert words_seen == {"alpha", "beta", "gamma"}

    def test_single_sheet_when_under_threshold(self, db_session: Session):
        """默认 ROWS_PER_SHEET=20000，3 个义项 → 单 sheet（命名仍为 _01）。"""
        from openpyxl import load_workbook

        for w in ("kappa", "lambda", "mu"):
            word = _make_word(db_session, w)
            meaning = _make_meaning(db_session, word, "n.", f"def-{w}")
            _make_content(
                db_session, word, "chunk", f"chunk-{w}",
                meaning=meaning, qc_status=QcStatus.APPROVED.value,
            )

        path = ExportService().export_to_excel(db_session)
        wb = load_workbook(path)

        assert wb.sheetnames == ["词表导出_01"]
        assert wb["词表导出_01"].max_row == 4  # 1 表头 + 3 数据
