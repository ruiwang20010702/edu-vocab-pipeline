"""生产记录（按 Package 聚合 AI 用量）集成测试."""

import pytest
from sqlalchemy.orm import Session

from vocab_qc.core.models.package_layer import Package, PackageWord
from vocab_qc.core.models.quality_layer import AiUsageLog
from vocab_qc.core.models.data_layer import Word
from vocab_qc.core.services.stats_service import get_production_records


def _make_word(session: Session, word: str) -> Word:
    w = Word(word=word)
    session.add(w)
    session.flush()
    return w


def _make_package(session: Session, name: str, total_words: int = 0, status: str = "completed") -> Package:
    pkg = Package(name=name, total_words=total_words, status=status)
    session.add(pkg)
    session.flush()
    return pkg


class TestGetProductionRecords:
    """get_production_records 按 Package 聚合测试。"""

    def test_empty(self, db_session: Session):
        """无 Package 时返回空列表。"""
        result = get_production_records(db_session)
        assert result["records"] == []
        assert result["totals"]["words"] == 0

    def test_pending_excluded(self, db_session: Session):
        """status=pending 的 Package 不包含在结果中。"""
        _make_package(db_session, "pending-pkg", status="pending")
        result = get_production_records(db_session)
        assert result["records"] == []

    def test_with_data(self, db_session: Session):
        """构造 Package + AiUsageLog，验证聚合结果正确。"""
        pkg = _make_package(db_session, "batch-1", total_words=10)
        w = _make_word(db_session, "hello")
        db_session.add(PackageWord(package_id=pkg.id, word_id=w.id))

        # Gemini 调用
        db_session.add(AiUsageLog(
            phase="generation",
            dimension="chunk",
            ai_model="gemini-3-flash-preview",
            prompt_tokens=630,
            completion_tokens=100,
            total_tokens=730,
            word_id=w.id,
            package_id=pkg.id,
        ))
        # GPT 调用
        db_session.add(AiUsageLog(
            phase="generation",
            dimension="sentence",
            ai_model="gpt-5.2",
            prompt_tokens=1120,
            completion_tokens=200,
            total_tokens=1320,
            word_id=w.id,
            package_id=pkg.id,
        ))
        # QC L2（Gemini）
        db_session.add(AiUsageLog(
            phase="qc_layer2",
            dimension=None,
            ai_model="gemini-3-flash-preview",
            prompt_tokens=500,
            completion_tokens=80,
            total_tokens=580,
            package_id=pkg.id,
        ))
        db_session.flush()

        result = get_production_records(db_session)
        assert len(result["records"]) == 1

        rec = result["records"][0]
        assert rec["batch_id"] == pkg.id
        assert rec["batch_name"] == "batch-1"
        assert rec["word_count"] == 10
        assert rec["gemini_input_tokens"] == 630 + 500
        assert rec["gemini_output_tokens"] == 100 + 80
        assert rec["gpt_input_tokens"] == 1120
        assert rec["gpt_output_tokens"] == 200

    def test_model_classification(self, db_session: Session):
        """验证不同模型名正确分类为 gemini / gpt。"""
        pkg = _make_package(db_session, "model-test", total_words=5)

        models_gemini = ["gemini-3-flash-preview", "gemini-2.0-flash", "Gemini-2.5-Flash"]
        models_gpt = ["gpt-5.2", "gpt-4o-mini", "GPT-4o"]

        for model in models_gemini:
            db_session.add(AiUsageLog(
                phase="generation", dimension="chunk", ai_model=model,
                prompt_tokens=100, completion_tokens=10, total_tokens=110,
                package_id=pkg.id,
            ))
        for model in models_gpt:
            db_session.add(AiUsageLog(
                phase="generation", dimension="sentence", ai_model=model,
                prompt_tokens=200, completion_tokens=20, total_tokens=220,
                package_id=pkg.id,
            ))
        db_session.flush()

        result = get_production_records(db_session)
        rec = result["records"][0]
        assert rec["gemini_input_tokens"] == 300  # 3 × 100
        assert rec["gemini_output_tokens"] == 30   # 3 × 10
        assert rec["gpt_input_tokens"] == 600       # 3 × 200
        assert rec["gpt_output_tokens"] == 60        # 3 × 20

    def test_multiple_packages(self, db_session: Session):
        """多个 Package 各自独立聚合。"""
        pkg1 = _make_package(db_session, "pkg-a", total_words=5)
        pkg2 = _make_package(db_session, "pkg-b", total_words=8)

        db_session.add(AiUsageLog(
            phase="generation", dimension="chunk", ai_model="gemini-3-flash",
            prompt_tokens=100, completion_tokens=10, total_tokens=110,
            package_id=pkg1.id,
        ))
        db_session.add(AiUsageLog(
            phase="generation", dimension="chunk", ai_model="gemini-3-flash",
            prompt_tokens=200, completion_tokens=20, total_tokens=220,
            package_id=pkg2.id,
        ))
        db_session.flush()

        result = get_production_records(db_session)
        assert len(result["records"]) == 2
        assert result["totals"]["words"] == 13
        assert result["totals"]["gemini_input"] == 300

    def test_package_no_usage(self, db_session: Session):
        """Package 存在但无 AI 调用时，token 为 0。"""
        _make_package(db_session, "empty-pkg", total_words=3)

        result = get_production_records(db_session)
        assert len(result["records"]) == 1
        rec = result["records"][0]
        assert rec["gemini_input_tokens"] == 0
        assert rec["gpt_input_tokens"] == 0


class TestExportProductionRecordsXlsx:
    """export_production_records_xlsx 导出测试。"""

    def test_export_xlsx(self, db_session: Session):
        """验证导出的 xlsx 格式正确。"""
        from openpyxl import load_workbook
        import io

        from vocab_qc.core.services.stats_service import export_production_records_xlsx

        pkg = _make_package(db_session, "export-test", total_words=5)
        db_session.add(AiUsageLog(
            phase="generation", dimension="chunk", ai_model="gemini-3-flash",
            prompt_tokens=100, completion_tokens=10, total_tokens=110,
            package_id=pkg.id,
        ))
        db_session.flush()

        buf = export_production_records_xlsx(db_session)
        assert isinstance(buf, io.BytesIO)

        wb = load_workbook(buf)
        ws = wb.active
        assert ws.title == "生产记录"

        # 表头
        assert ws.cell(row=1, column=1).value == "批次号"
        assert ws.cell(row=1, column=6).value == "Gemini\nInput Tokens"

        # 数据行
        assert ws.cell(row=2, column=2).value == "export-test"
        assert ws.cell(row=2, column=3).value == 5
        assert ws.cell(row=2, column=6).value == 100
