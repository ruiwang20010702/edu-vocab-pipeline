"""一次性脚本：导出 3 个助记维度（排除考试应用）各 50 条精选已审内容到 Excel。

精选口径：人工从随机已审池中按"公式拆解正确 + 口诀联想贴切自然、不牵强"挑选。
ID 清单为固定快照，保证可复现。
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from vocab_qc.core.db import SyncSessionLocal

SHEETS: list[tuple[str, str, list[int]]] = [
    (
        "词根词缀",
        "mnemonic_root_affix",
        [
            101202, 62724, 111336, 14408, 57826, 167300, 177916, 216744, 172664, 94188,
            121571, 6495, 40001, 151843, 98916, 76770, 172874, 96414, 96870, 79098,
            150871, 43565, 133306, 154895, 156545, 85240, 235026, 147121, 71448, 236424,
            78414, 191331, 35795, 54730, 254644, 124325, 51118, 221130, 230612, 250384,
            128381, 31937, 16304, 98676, 139420, 29741, 229928, 41819, 62064, 220278,
        ],
    ),
    (
        "词中词",
        "mnemonic_word_in_word",
        [
            74018, 182430, 161347, 67334, 239804, 12193, 231838, 230086, 5281, 103039,
            7183, 225664, 186138, 72884, 218120, 132822, 188795, 151635, 103051, 33721,
            210110, 125310, 175636, 178182, 262794, 177846, 32965, 131772, 230110, 228220,
            171958, 168946, 89709, 74240, 58590, 111823, 241790, 145929, 54102, 74996,
            241574, 226774, 31261, 124050, 34411, 35179, 72008, 65438, 91185, 138954,
        ],
    ),
    (
        "音义联想",
        "mnemonic_sound_meaning",
        [
            34104, 61087, 138977, 127291, 172065, 75109, 167067, 189760, 127837, 109418,
            72211, 68701, 159234, 156828, 42684, 132503, 115052, 69031, 239167, 219463,
            11412, 146822, 186845, 148862, 52349, 182309, 224691, 154764, 135545, 95690,
            16764, 150746, 140945, 25487, 66745, 32508, 100916, 148472, 146420, 132905,
            116750, 218845, 174537, 207973, 120205, 191428, 17268, 111818, 46950, 140165,
        ],
    ),
]

HEADERS = ["序号", "单词", "词性", "释义", "公式", "口诀", "老师话术"]
WIDTHS = [6, 18, 8, 28, 32, 30, 70]


def fetch_rows(session, ids: list[int]) -> dict[int, dict]:
    rows = session.execute(
        text(
            """
            select c.id, w.word, m.pos, m.definition, c.content
            from content_items c
            join words w on w.id = c.word_id
            left join meanings m on m.id = c.meaning_id
            where c.id = any(:ids)
            """
        ),
        {"ids": ids},
    ).fetchall()
    out: dict[int, dict] = {}
    for r in rows:
        data = json.loads(r[4])
        out[r[0]] = {
            "word": r[1],
            "pos": r[2] or "",
            "definition": r[3] or "",
            "formula": data.get("formula", ""),
            "chant": data.get("chant", ""),
            "exam_sentence": data.get("exam_sentence", ""),
            "exam_sentence_translation": data.get("exam_sentence_translation", ""),
            "script": data.get("script", ""),
        }
    return out


# 第 4 维度「考试应用（考点拓展）」结构多两列：实战例句 + 例句翻译。
# 全文逐条核验（搭配真伪/例句语法/翻译准确/话术无误），优先高频经典考点。
EXAM_SHEET = (
    "考试应用",
    "mnemonic_exam_app",
    [
        1766, 4520, 8732, 16977, 26830, 38870, 40472, 46832, 59035, 62079,
        71721, 75543, 95235, 98577, 100833, 120080, 126608, 129008, 132991, 146296,
        160598, 172061, 178039, 189966, 192180, 207315, 207693, 213867, 214551, 224207,
        243429, 246727, 251155, 254275, 33086, 2478, 4514, 18291, 20121, 24915,
        26854, 39326, 62793, 84260, 86162, 201525, 216765, 257089, 177979, 124484,
    ],
)
EXAM_HEADERS = ["序号", "单词", "词性", "释义", "考点公式", "考点逻辑", "实战例句", "例句翻译", "老师话术"]
EXAM_WIDTHS = [6, 18, 8, 24, 30, 26, 46, 36, 64]


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    with SyncSessionLocal() as session:
        for sheet_name, dimension, ids in SHEETS:
            data = fetch_rows(session, ids)
            ws = wb.create_sheet(title=sheet_name)
            for col, h in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col)].width = WIDTHS[col - 1]
            ws.freeze_panes = "A2"
            for i, cid in enumerate(ids, 1):
                d = data.get(cid)
                if not d:
                    print(f"WARN: id {cid} not found in {dimension}")
                    continue
                vals = [i, d["word"], d["pos"], d["definition"], d["formula"], d["chant"], d["script"]]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=i + 1, column=col, value=v)
                    cell.alignment = wrap_top
            print(f"{sheet_name}: 写入 {len(ids)} 条")

        # 第 4 sheet：考试应用（考点拓展），9 列
        sheet_name, _dim, ids = EXAM_SHEET
        data = fetch_rows(session, ids)
        ws = wb.create_sheet(title=sheet_name)
        for col, h in enumerate(EXAM_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = EXAM_WIDTHS[col - 1]
        ws.freeze_panes = "A2"
        for i, cid in enumerate(ids, 1):
            d = data.get(cid)
            if not d:
                print(f"WARN: id {cid} not found in exam_app")
                continue
            vals = [
                i, d["word"], d["pos"], d["definition"], d["formula"], d["chant"],
                d["exam_sentence"], d["exam_sentence_translation"], d["script"],
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=i + 1, column=col, value=v).alignment = wrap_top
        print(f"{sheet_name}: 写入 {len(ids)} 条")

    out_path = "助记精选样例_4维度x50.xlsx"
    wb.save(out_path)
    print(f"已保存: {out_path}")


if __name__ == "__main__":
    main()
