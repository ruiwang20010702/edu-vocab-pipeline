"""
S9 词汇生产 AI 成本估算 Excel 生成器

生成包含 4 个 Sheet 的工作簿：
1. 参数配置 - 模型定价、维度映射、调整因子、总词数输入
2. 预算估算 - 3 场景自动计算 + 最差情况
3. 生产记录 - 批次级实际数据录入模板
4. 汇总分析 - 预算 vs 实际对比 + 关键指标
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from pathlib import Path


# ── 样式常量 ──────────────────────────────────────────────

_TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True)
_HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
_SECTION_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
_NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
_SMALL_FONT = Font(name="Microsoft YaHei", size=9, color="666666")
_INPUT_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="0066CC")

_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_GEMINI_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_GPT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_WORST_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")

# 数字格式
_FMT_CNY = '¥#,##0.00'
_FMT_CNY4 = '¥#,##0.0000'
_FMT_INT = '#,##0'
_FMT_PCT = '0.0%'
_FMT_TOKEN = '#,##0'


def _apply_header_row(ws, row, cols, headers):
    """写入表头行并设置样式。"""
    for i, h in enumerate(headers, start=cols[0] if isinstance(cols, range) else 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _apply_section_label(ws, row, col, text, merge_end=None):
    """写入分区标签。"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.border = _THIN_BORDER
    if merge_end:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=merge_end
        )
        for c in range(col, merge_end + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
            ws.cell(row=row, column=c).border = _THIN_BORDER


def _write_row(ws, row, data, start_col=1, font=None, fill=None, fmt=None):
    """写入一行数据。"""
    f = font or _NORMAL_FONT
    for i, val in enumerate(data, start=start_col):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = f
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
        if fill:
            cell.fill = fill
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt


def _set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ── 维度数据 ──────────────────────────────────────────────

# (维度名, 模型标识, 调用粒度, input_tokens, output_tokens, prompt字节数)
GEN_DIMENSIONS = [
    ("syllable（音节）", "Gemini", "按单词", 950, 50, 2192),
    ("chunk（语块）", "Gemini", "按义项", 630, 100, 1460),
    ("sentence（例句）", "GPT-5.2", "按义项", 1120, 200, 2580),
    ("mnemonic_root_affix（词根词缀）", "Gemini", "按义项", 2330, 400, 5370),
    ("mnemonic_word_in_word（词中词）", "Gemini", "按义项", 1830, 350, 4214),
    ("mnemonic_sound_meaning（音义联想）", "Gemini", "按义项", 1750, 350, 4034),
    ("mnemonic_exam_app（考试应用）", "Gemini", "按义项", 1910, 350, 4406),
]

QC_DIMENSIONS = [
    ("syllable_unified", "Gemini", "按单词", 600, 80),
    ("chunk_ai", "Gemini", "按义项", 500, 80),
    ("sentence_ai", "Gemini", "按义项", 1200, 150),
    ("mnemonic_root_affix_ai", "Gemini", "按义项", 900, 100),
    ("mnemonic_word_in_word_ai", "Gemini", "按义项", 900, 100),
    ("mnemonic_sound_meaning_ai", "Gemini", "按义项", 900, 100),
    ("mnemonic_exam_app_ai", "Gemini", "按义项", 900, 100),
]


# ── Sheet 1: 参数配置 ────────────────────────────────────

def _build_params_sheet(ws):
    ws.title = "参数配置"
    _set_col_widths(ws, {
        "A": 4, "B": 30, "C": 22, "D": 22, "E": 18,
        "F": 16, "G": 16, "H": 16,
    })

    # ── 标题 ──
    r = 1
    ws.cell(row=r, column=2, value="S9 词汇生产 AI 成本 — 参数配置").font = _TITLE_FONT
    ws.merge_cells("B1:F1")

    # ── 总词数输入 ──
    r = 3
    _apply_section_label(ws, r, 2, "核心输入", merge_end=3)
    r = 4
    ws.cell(row=r, column=2, value="总词数").font = _NORMAL_FONT
    ws.cell(row=r, column=2).border = _THIN_BORDER
    input_cell = ws.cell(row=r, column=3, value=500)
    input_cell.font = _INPUT_FONT
    input_cell.fill = _INPUT_FILL
    input_cell.border = _THIN_BORDER
    input_cell.alignment = _CENTER
    input_cell.number_format = _FMT_INT
    ws.cell(row=r, column=4, value="← 修改此处，所有计算自动更新").font = _SMALL_FONT

    # ── 模型定价 ──
    r = 6
    _apply_section_label(ws, r, 2, "模型定价（¥/百万 token · AI Gateway 计价）", merge_end=6)
    r = 7
    _apply_header_row(ws, r, range(2, 7), ["模型", "TEXT INPUT", "TEXT OUTPUT", "数据来源", "备注"])
    r = 8
    _write_row(ws, r, ["gemini-3-flash-preview", 3.65, 21.90, "model_id=64", "6个生成维度+全部QC"], start_col=2, fill=_GEMINI_FILL)
    ws.cell(row=r, column=3).number_format = '0.00'
    ws.cell(row=r, column=4).number_format = '0.00'
    r = 9
    _write_row(ws, r, ["gpt-5.2", 12.78, 102.20, "model_id=78", "仅例句生成"], start_col=2, fill=_GPT_FILL)
    ws.cell(row=r, column=3).number_format = '0.00'
    ws.cell(row=r, column=4).number_format = '0.00'

    # ── 维度→模型映射 + Token 预估 ──
    r = 11
    _apply_section_label(ws, r, 2, "生成维度 → Token 预估（单义词基准）", merge_end=8)
    r = 12
    _apply_header_row(ws, r, range(2, 9), ["维度", "模型", "调用粒度", "Input Tokens", "Output Tokens", "Prompt(字节)", "备注"])
    for i, dim in enumerate(GEN_DIMENSIONS):
        r = 13 + i
        fill = _GPT_FILL if dim[1] == "GPT-5.2" else _GEMINI_FILL
        _write_row(ws, r, list(dim), start_col=2, fill=fill)
    r = 13 + len(GEN_DIMENSIONS)
    _write_row(ws, r, [
        "生成合计", "混合", "—",
        sum(d[3] for d in GEN_DIMENSIONS),
        sum(d[4] for d in GEN_DIMENSIONS),
        sum(d[5] for d in GEN_DIMENSIONS),
        "7次AI调用/词",
    ], start_col=2, fill=_TOTAL_FILL, font=_SECTION_FONT)

    r += 2
    _apply_section_label(ws, r, 2, "质检 Layer2 维度 → Token 预估（全 Gemini）", merge_end=8)
    r += 1
    _apply_header_row(ws, r, range(2, 8), ["维度", "模型", "调用粒度", "Input Tokens", "Output Tokens", "备注"])
    qc_start = r + 1
    for i, dim in enumerate(QC_DIMENSIONS):
        r = qc_start + i
        _write_row(ws, r, list(dim), start_col=2, fill=_GEMINI_FILL)
    r = qc_start + len(QC_DIMENSIONS)
    _write_row(ws, r, [
        "质检L2合计", "Gemini", "—",
        sum(d[3] for d in QC_DIMENSIONS),
        sum(d[4] for d in QC_DIMENSIONS),
        "7次AI调用/词",
    ], start_col=2, fill=_TOTAL_FILL, font=_SECTION_FONT)

    # ── 调整因子 ──
    r += 2
    factor_section_row = r
    _apply_section_label(ws, r, 2, "调整因子", merge_end=7)
    r += 1
    _apply_header_row(ws, r, range(2, 8), ["因子", "说明", "乐观", "中等", "悲观", "备注"])
    factors = [
        ("平均义项数", "多义词增加按义项维度调用", 1.0, 1.3, 1.8, "中小学词汇多义词约30%"),
        ("助记有效率", "valid=false不进QC L2", 0.60, 0.75, 0.90, "部分词不适合某类助记"),
        ("L1通过率", "L1未通过不进L2", 0.90, 0.95, 0.98, ""),
        ("审核重生成率", "reject后重新生成", 0.02, 0.05, 0.10, "最多3次"),
        ("API重试率", "网络/限流导致重试", 0.01, 0.03, 0.05, ""),
    ]
    factor_data_start = r + 1
    for i, f in enumerate(factors):
        r = factor_data_start + i
        _write_row(ws, r, list(f), start_col=2)
        # 百分比格式（除了平均义项数）
        if i > 0:
            for c in [4, 5, 6]:
                ws.cell(row=r, column=c).number_format = _FMT_PCT

    # ── Layer1 说明 ──
    r += 2
    ws.cell(row=r, column=2, value="说明：Layer1 质检为 22 条纯算法规则，零 AI 成本").font = _SMALL_FONT

    return factor_data_start


# ── Sheet 2: 预算估算 ────────────────────────────────────

def _build_budget_sheet(wb, factor_start_row):
    ws = wb.create_sheet("预算估算")
    _set_col_widths(ws, {
        "A": 4, "B": 32, "C": 16, "D": 18, "E": 18,
        "F": 16, "G": 16, "H": 16, "I": 16,
    })

    # 引用参数配置 Sheet
    P = "参数配置"  # sheet name for refs
    word_ref = f"'{P}'!C4"  # 总词数

    # Gemini 定价引用
    gemini_in_ref = f"'{P}'!C8"   # 3.65
    gemini_out_ref = f"'{P}'!D8"  # 21.90
    gpt_in_ref = f"'{P}'!C9"     # 12.78
    gpt_out_ref = f"'{P}'!D9"    # 102.20

    # 调整因子引用：乐观=D列, 中等=E列, 悲观=F列
    # factor_start_row 对应：义项数, 助记有效率, L1通过率, 重生成率, API重试率
    fr = factor_start_row  # row of 平均义项数

    r = 1
    ws.cell(row=r, column=2, value="S9 词汇生产 AI 成本 — 预算估算").font = _TITLE_FONT
    ws.merge_cells("B1:G1")

    r = 2
    ws.cell(row=r, column=2).font = _NORMAL_FONT
    # 动态显示总词数
    ws.cell(row=r, column=2, value=f"总词数：").font = _NORMAL_FONT
    ws.cell(row=r, column=3).font = _INPUT_FONT
    ws.cell(row=r, column=3).number_format = _FMT_INT
    ws.cell(row=r, column=3, value=f"={word_ref}")
    ws.cell(row=r, column=4, value="（修改「参数配置」Sheet C4 单元格）").font = _SMALL_FONT

    # ── 三场景计算 ──
    scenarios = [
        ("乐观", "D", _GEMINI_FILL),
        ("中等（推荐参考）", "E", _TOTAL_FILL),
        ("悲观", "F", _WORST_FILL),
    ]

    r = 4
    for s_idx, (s_name, col_letter, fill) in enumerate(scenarios):
        _apply_section_label(ws, r, 2, f"场景：{s_name}", merge_end=9)
        r += 1
        _apply_header_row(ws, r, range(2, 10), [
            "项目", "模型", "调用次数", "Input Tokens", "Output Tokens",
            "Input成本(¥)", "Output成本(¥)", "总成本(¥)",
        ])

        data_start = r + 1
        sense_ref = f"'{P}'!{col_letter}{fr}"      # 义项数
        mnem_eff = f"'{P}'!{col_letter}{fr+1}"     # 助记有效率
        l1_pass = f"'{P}'!{col_letter}{fr+2}"      # L1通过率
        regen = f"'{P}'!{col_letter}{fr+3}"         # 重生成率
        retry = f"'{P}'!{col_letter}{fr+4}"         # API重试率

        # -- 生成阶段 Gemini (6维度) --
        r = data_start
        ws.cell(row=r, column=2, value="生成 Gemini（6维度）").font = _NORMAL_FONT
        ws.cell(row=r, column=3, value="Gemini").font = _NORMAL_FONT
        # syllable 按单词(1次) + 其余5个Gemini维度按义项
        # Gemini input per word = 950 + (630+2330+1830+1750+1910)*sense
        # Gemini output per word = 50 + (100+400+350+350+350)*sense
        gemini_gen_in = f"({word_ref}*(950+(630+2330+1830+1750+1910)*{sense_ref}))"
        gemini_gen_out = f"({word_ref}*(50+(100+400+350+350+350)*{sense_ref}))"
        gemini_gen_calls = f"({word_ref}*(1+5*{sense_ref}))"
        ws.cell(row=r, column=4, value=f"={gemini_gen_calls}").number_format = _FMT_INT
        ws.cell(row=r, column=5, value=f"={gemini_gen_in}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=6, value=f"={gemini_gen_out}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=7, value=f"={gemini_gen_in}*{gemini_in_ref}/1000000").number_format = _FMT_CNY4
        ws.cell(row=r, column=8, value=f"={gemini_gen_out}*{gemini_out_ref}/1000000").number_format = _FMT_CNY4
        ws.cell(row=r, column=9, value=f"={get_column_letter(7)}{r}+{get_column_letter(8)}{r}").number_format = _FMT_CNY
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = _THIN_BORDER
            ws.cell(row=r, column=c).alignment = _CENTER
            ws.cell(row=r, column=c).fill = _GEMINI_FILL

        # -- 生成阶段 GPT (例句) --
        r += 1
        ws.cell(row=r, column=2, value="生成 GPT-5.2（例句）").font = _NORMAL_FONT
        ws.cell(row=r, column=3, value="GPT-5.2").font = _NORMAL_FONT
        gpt_gen_in = f"({word_ref}*1120*{sense_ref})"
        gpt_gen_out = f"({word_ref}*200*{sense_ref})"
        gpt_gen_calls = f"({word_ref}*{sense_ref})"
        ws.cell(row=r, column=4, value=f"={gpt_gen_calls}").number_format = _FMT_INT
        ws.cell(row=r, column=5, value=f"={gpt_gen_in}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=6, value=f"={gpt_gen_out}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=7, value=f"={gpt_gen_in}*{gpt_in_ref}/1000000").number_format = _FMT_CNY4
        ws.cell(row=r, column=8, value=f"={gpt_gen_out}*{gpt_out_ref}/1000000").number_format = _FMT_CNY4
        ws.cell(row=r, column=9, value=f"={get_column_letter(7)}{r}+{get_column_letter(8)}{r}").number_format = _FMT_CNY
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = _THIN_BORDER
            ws.cell(row=r, column=c).alignment = _CENTER
            ws.cell(row=r, column=c).fill = _GPT_FILL

        # -- 质检 L2 (全 Gemini) --
        r += 1
        ws.cell(row=r, column=2, value="质检 L2（全 Gemini）").font = _NORMAL_FONT
        ws.cell(row=r, column=3, value="Gemini").font = _NORMAL_FONT
        # syllable_unified 按单词(1次) + 其余6个按义项，助记4个受有效率影响
        # 非助记按义项：chunk(500in,80out) + sentence(1200in,150out) = 1700in,230out
        # 助记按义项×有效率：4×(900in,100out) = 3600in,400out
        # L1通过率影响全部
        qc_in = f"({word_ref}*(600+(1700*{sense_ref}+3600*{sense_ref}*{mnem_eff})*{l1_pass}))"
        qc_out = f"({word_ref}*(80+(230*{sense_ref}+400*{sense_ref}*{mnem_eff})*{l1_pass}))"
        qc_calls = f"({word_ref}*(1+(2*{sense_ref}+4*{sense_ref}*{mnem_eff})*{l1_pass}))"
        ws.cell(row=r, column=4, value=f"={qc_calls}").number_format = _FMT_INT
        ws.cell(row=r, column=5, value=f"={qc_in}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=6, value=f"={qc_out}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=7, value=f"={qc_in}*{gemini_in_ref}/1000000").number_format = _FMT_CNY4
        ws.cell(row=r, column=8, value=f"={qc_out}*{gemini_out_ref}/1000000").number_format = _FMT_CNY4
        ws.cell(row=r, column=9, value=f"={get_column_letter(7)}{r}+{get_column_letter(8)}{r}").number_format = _FMT_CNY
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = _THIN_BORDER
            ws.cell(row=r, column=c).alignment = _CENTER
            ws.cell(row=r, column=c).fill = _GEMINI_FILL

        # -- 重试+重生成 --
        r += 1
        ws.cell(row=r, column=2, value="重试+重生成").font = _NORMAL_FONT
        ws.cell(row=r, column=3, value="混合").font = _NORMAL_FONT
        # 综合调整 = (重生成率 + API重试率) × 前三行合计
        base_cost_range = f"I{data_start}:I{r-1}"
        base_calls_range = f"D{data_start}:D{r-1}"
        base_in_range = f"E{data_start}:E{r-1}"
        base_out_range = f"F{data_start}:F{r-1}"
        adj_factor = f"({regen}+{retry})"
        ws.cell(row=r, column=4, value=f"=SUM({base_calls_range})*{adj_factor}").number_format = _FMT_INT
        ws.cell(row=r, column=5, value=f"=SUM({base_in_range})*{adj_factor}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=6, value=f"=SUM({base_out_range})*{adj_factor}").number_format = _FMT_TOKEN
        ws.cell(row=r, column=7, value="—").alignment = _CENTER
        ws.cell(row=r, column=8, value="—").alignment = _CENTER
        ws.cell(row=r, column=9, value=f"=SUM({base_cost_range})*{adj_factor}").number_format = _FMT_CNY
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = _THIN_BORDER
            ws.cell(row=r, column=c).alignment = _CENTER

        # -- 合计行 --
        r += 1
        total_row = r
        ws.cell(row=r, column=2, value=f"▶ {s_name} 合计").font = _SECTION_FONT
        ws.cell(row=r, column=4, value=f"=SUM(D{data_start}:D{r-1})").number_format = _FMT_INT
        ws.cell(row=r, column=5, value=f"=SUM(E{data_start}:E{r-1})").number_format = _FMT_TOKEN
        ws.cell(row=r, column=6, value=f"=SUM(F{data_start}:F{r-1})").number_format = _FMT_TOKEN
        ws.cell(row=r, column=9, value=f"=SUM(I{data_start}:I{r-1})").number_format = _FMT_CNY
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = _THIN_BORDER
            ws.cell(row=r, column=c).alignment = _CENTER
            ws.cell(row=r, column=c).fill = fill

        # -- 每词均价 --
        r += 1
        ws.cell(row=r, column=2, value="每词均价").font = _NORMAL_FONT
        ws.cell(row=r, column=2).border = _THIN_BORDER
        ws.cell(row=r, column=9, value=f"=I{total_row}/{word_ref}").number_format = _FMT_CNY4
        ws.cell(row=r, column=9).border = _THIN_BORDER
        ws.cell(row=r, column=9).alignment = _CENTER
        ws.cell(row=r, column=9).fill = fill

        # 记录中等场景合计行位置
        if s_name.startswith("中等"):
            ws._medium_total_row = total_row

        r += 2  # 间隔

    # ── 最差情况 ──
    _apply_section_label(ws, r, 2, "极端场景：全部走满 3 轮重生成（×4 轮生产+质检）", merge_end=9)
    r += 1
    # 悲观基础 × 4 轮
    worst_base = total_row  # 最后一个场景（悲观）的合计行
    ws.cell(row=r, column=2, value="悲观基础成本 × 4 轮").font = _NORMAL_FONT
    ws.cell(row=r, column=9, value=f"=I{worst_base}*4").number_format = _FMT_CNY
    for c in [2, 9]:
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER
        ws.cell(row=r, column=c).fill = _WORST_FILL
    worst_total_row = r
    r += 1
    ws.cell(row=r, column=2, value="最差每词均价").font = _NORMAL_FONT
    ws.cell(row=r, column=9, value=f"=I{worst_total_row}/{word_ref}").number_format = _FMT_CNY4
    for c in [2, 9]:
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER
        ws.cell(row=r, column=c).fill = _WORST_FILL

    return ws


# ── Sheet 3: 生产记录 ────────────────────────────────────

def _build_record_sheet(wb):
    ws = wb.create_sheet("生产记录")
    _set_col_widths(ws, {
        "A": 4, "B": 8, "C": 16, "D": 8, "E": 18, "F": 18,
        "G": 18, "H": 18, "I": 18, "J": 18,
        "K": 16, "L": 16, "M": 16, "N": 20,
    })

    r = 1
    ws.cell(row=r, column=2, value="S9 词汇生产 — 实际成本记录").font = _TITLE_FONT
    ws.merge_cells("B1:H1")

    r = 2
    ws.cell(row=r, column=2, value="说明：每批生产完成后填入实际数据，成本列自动计算").font = _SMALL_FONT

    r = 3
    headers = [
        "批次号", "批次名称", "词数",
        "开始时间", "结束时间",
        "Gemini\nInput Tokens", "Gemini\nOutput Tokens",
        "GPT\nInput Tokens", "GPT\nOutput Tokens",
        "Gemini成本(¥)", "GPT成本(¥)",
        "总成本(¥)", "备注",
    ]
    _apply_header_row(ws, r, range(2, 15), headers)
    ws.row_dimensions[r].height = 36

    # Gemini/GPT 定价引用
    gemini_in_ref = "'参数配置'!C8"
    gemini_out_ref = "'参数配置'!D8"
    gpt_in_ref = "'参数配置'!C9"
    gpt_out_ref = "'参数配置'!D9"

    # 预留 50 行数据区
    data_start = 4
    data_end = 53
    for row in range(data_start, data_end + 1):
        ws.cell(row=row, column=2, value=row - data_start + 1).number_format = '0'
        for c in range(2, 15):
            ws.cell(row=row, column=c).border = _THIN_BORDER
            ws.cell(row=row, column=c).alignment = _CENTER
            ws.cell(row=row, column=c).font = _NORMAL_FONT

        # Token 列数字格式
        for c in [7, 8, 9, 10]:
            ws.cell(row=row, column=c).number_format = _FMT_TOKEN

        # 自动计算成本（当有 token 数据时）
        # Gemini 成本 = (G input × 定价 + H output × 定价) / 1000000
        g_col = get_column_letter(7)   # Gemini Input
        h_col = get_column_letter(8)   # Gemini Output
        i_col = get_column_letter(9)   # GPT Input
        j_col = get_column_letter(10)  # GPT Output

        # K列: Gemini成本
        ws.cell(row=row, column=11,
                value=f'=IF(OR({g_col}{row}<>"",{h_col}{row}<>""),({g_col}{row}*{gemini_in_ref}+{h_col}{row}*{gemini_out_ref})/1000000,"")').number_format = _FMT_CNY
        # L列: GPT成本
        ws.cell(row=row, column=12,
                value=f'=IF(OR({i_col}{row}<>"",{j_col}{row}<>""),({i_col}{row}*{gpt_in_ref}+{j_col}{row}*{gpt_out_ref})/1000000,"")').number_format = _FMT_CNY
        # M列: 总成本
        k_col = get_column_letter(11)
        l_col = get_column_letter(12)
        ws.cell(row=row, column=13,
                value=f'=IF(OR({k_col}{row}<>"",{l_col}{row}<>""),{k_col}{row}+{l_col}{row},"")').number_format = _FMT_CNY

    # ── 汇总行 ──
    r = data_end + 1
    ws.cell(row=r, column=2, value="汇总").font = _SECTION_FONT
    ws.merge_cells(f"B{r}:C{r}")
    # 词数合计
    ws.cell(row=r, column=4, value=f"=SUM(D{data_start}:D{data_end})").number_format = _FMT_INT
    ws.cell(row=r, column=4).font = _SECTION_FONT
    # Token 合计
    for c in [7, 8, 9, 10]:
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c, value=f"=SUM({col_l}{data_start}:{col_l}{data_end})").number_format = _FMT_TOKEN
        ws.cell(row=r, column=c).font = _SECTION_FONT
    # 成本合计
    for c in [11, 12, 13]:
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c, value=f"=SUM({col_l}{data_start}:{col_l}{data_end})").number_format = _FMT_CNY
        ws.cell(row=r, column=c).font = _SECTION_FONT
    for c in range(2, 15):
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER
        ws.cell(row=r, column=c).fill = _TOTAL_FILL

    # ── 每词均价 ──
    r += 1
    ws.cell(row=r, column=2, value="每词均价").font = _NORMAL_FONT
    ws.cell(row=r, column=2).border = _THIN_BORDER
    prev = r - 1
    ws.cell(row=r, column=13,
            value=f'=IF(D{prev}>0,M{prev}/D{prev},"")').number_format = _FMT_CNY4
    ws.cell(row=r, column=13).border = _THIN_BORDER
    ws.cell(row=r, column=13).alignment = _CENTER
    ws.cell(row=r, column=13).fill = _TOTAL_FILL

    return data_start, data_end


# ── Sheet 4: 汇总分析 ────────────────────────────────────

def _build_analysis_sheet(wb, record_data_start, record_data_end):
    ws = wb.create_sheet("汇总分析")
    _set_col_widths(ws, {
        "A": 4, "B": 28, "C": 20, "D": 20, "E": 20, "F": 20, "G": 16,
    })

    record_summary_row = record_data_end + 1  # 生产记录汇总行
    word_ref = "'参数配置'!C4"

    r = 1
    ws.cell(row=r, column=2, value="S9 词汇生产 AI 成本 — 汇总分析").font = _TITLE_FONT
    ws.merge_cells("B1:F1")

    # ── 关键指标卡 ──
    r = 3
    _apply_section_label(ws, r, 2, "关键指标", merge_end=6)
    r = 4
    metrics = [
        ("计划总词数", f"={word_ref}", _FMT_INT),
        ("已完成词数", f"='生产记录'!D{record_summary_row}", _FMT_INT),
        ("完成进度", f"=IF({word_ref}>0,'生产记录'!D{record_summary_row}/{word_ref},0)", _FMT_PCT),
    ]
    _apply_header_row(ws, r, range(2, 5), ["指标", "数值", "格式说明"])
    for i, (name, formula, fmt) in enumerate(metrics):
        r = 5 + i
        ws.cell(row=r, column=2, value=name).font = _NORMAL_FONT
        ws.cell(row=r, column=2).border = _THIN_BORDER
        cell = ws.cell(row=r, column=3, value=formula)
        cell.number_format = fmt
        cell.font = _INPUT_FONT
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER

    # ── 预算 vs 实际 ──
    r = 9
    _apply_section_label(ws, r, 2, "预算 vs 实际对比", merge_end=6)
    r = 10
    _apply_header_row(ws, r, range(2, 7), ["项目", "预算（中等）", "实际", "偏差", "偏差率"])

    # 总成本
    r = 11
    # 找中等场景合计 - 通过 ws 的属性获取
    budget_ws = wb["预算估算"]
    medium_total = getattr(budget_ws, '_medium_total_row', 14)  # fallback
    ws.cell(row=r, column=2, value="总成本").font = _NORMAL_FONT
    ws.cell(row=r, column=3, value=f"='预算估算'!I{medium_total}").number_format = _FMT_CNY
    ws.cell(row=r, column=4, value=f"='生产记录'!M{record_summary_row}").number_format = _FMT_CNY
    ws.cell(row=r, column=5, value=f"=IF(D{r}<>\"\",D{r}-C{r},\"\")").number_format = _FMT_CNY
    ws.cell(row=r, column=6, value=f"=IF(AND(C{r}>0,D{r}<>\"\"),D{r}/C{r}-1,\"\")").number_format = _FMT_PCT
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER

    # 每词均价
    r = 12
    ws.cell(row=r, column=2, value="每词均价").font = _NORMAL_FONT
    ws.cell(row=r, column=3, value=f"=IF({word_ref}>0,'预算估算'!I{medium_total}/{word_ref},0)").number_format = _FMT_CNY4
    ws.cell(row=r, column=4, value=f"=IF('生产记录'!D{record_summary_row}>0,'生产记录'!M{record_summary_row}/'生产记录'!D{record_summary_row},\"\")").number_format = _FMT_CNY4
    ws.cell(row=r, column=5, value=f"=IF(D{r}<>\"\",D{r}-C{r},\"\")").number_format = _FMT_CNY4
    ws.cell(row=r, column=6, value=f"=IF(AND(C{r}>0,D{r}<>\"\"),D{r}/C{r}-1,\"\")").number_format = _FMT_PCT
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER

    # ── 成本构成 ──
    r = 14
    _apply_section_label(ws, r, 2, "实际成本构成（按模型）", merge_end=6)
    r = 15
    _apply_header_row(ws, r, range(2, 6), ["模型", "成本(¥)", "占比", "说明"])

    r = 16
    g_col = get_column_letter(11)  # 生产记录 K列 = Gemini成本
    l_col = get_column_letter(12)  # 生产记录 L列 = GPT成本
    ws.cell(row=r, column=2, value="Gemini").font = _NORMAL_FONT
    ws.cell(row=r, column=3, value=f"='生产记录'!{g_col}{record_summary_row}").number_format = _FMT_CNY
    ws.cell(row=r, column=4,
            value=f"=IF('生产记录'!M{record_summary_row}>0,C{r}/'生产记录'!M{record_summary_row},\"\")").number_format = _FMT_PCT
    ws.cell(row=r, column=5, value="6个生成维度 + 全部QC L2").font = _SMALL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER
        ws.cell(row=r, column=c).fill = _GEMINI_FILL

    r = 17
    ws.cell(row=r, column=2, value="GPT-5.2").font = _NORMAL_FONT
    ws.cell(row=r, column=3, value=f"='生产记录'!{l_col}{record_summary_row}").number_format = _FMT_CNY
    ws.cell(row=r, column=4,
            value=f"=IF('生产记录'!M{record_summary_row}>0,C{r}/'生产记录'!M{record_summary_row},\"\")").number_format = _FMT_PCT
    ws.cell(row=r, column=5, value="仅例句维度").font = _SMALL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = _THIN_BORDER
        ws.cell(row=r, column=c).alignment = _CENTER
        ws.cell(row=r, column=c).fill = _GPT_FILL

    # ── 批次明细统计 ──
    r = 19
    _apply_section_label(ws, r, 2, "批次成本明细（每词均价趋势）", merge_end=6)
    r = 20
    _apply_header_row(ws, r, range(2, 7), ["批次号", "词数", "总成本(¥)", "每词均价(¥)", "累计均价(¥)"])
    # 引用前 20 批
    for i in range(20):
        row_src = record_data_start + i
        r = 21 + i
        ws.cell(row=r, column=2, value=f"='生产记录'!B{row_src}").number_format = '0'
        ws.cell(row=r, column=3, value=f"='生产记录'!D{row_src}").number_format = _FMT_INT
        ws.cell(row=r, column=4, value=f"='生产记录'!M{row_src}").number_format = _FMT_CNY
        ws.cell(row=r, column=5,
                value=f"=IF('生产记录'!D{row_src}>0,'生产记录'!M{row_src}/'生产记录'!D{row_src},\"\")").number_format = _FMT_CNY4
        # 累计均价 = 到当前批次的总成本 / 总词数
        ws.cell(row=r, column=6,
                value=f"=IF(SUM('生产记录'!D{record_data_start}:D{row_src})>0,SUM('生产记录'!M{record_data_start}:M{row_src})/SUM('生产记录'!D{record_data_start}:D{row_src}),\"\")").number_format = _FMT_CNY4
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = _THIN_BORDER
            ws.cell(row=r, column=c).alignment = _CENTER

    return ws


# ── 主函数 ────────────────────────────────────────────────

def main():
    wb = Workbook()

    # Sheet 1
    ws_params = wb.active
    factor_start_row = _build_params_sheet(ws_params)

    # Sheet 2
    _build_budget_sheet(wb, factor_start_row)

    # Sheet 3
    data_start, data_end = _build_record_sheet(wb)

    # Sheet 4
    _build_analysis_sheet(wb, data_start, data_end)

    # 输出
    out_path = Path(__file__).resolve().parent.parent / "docs" / "S9词汇生产AI成本估算.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"已生成: {out_path}")
    print(f"文件大小: {out_path.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
