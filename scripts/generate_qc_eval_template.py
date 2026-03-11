"""生成质检模型评估模板：案例表 + 评估表."""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 样式 ──────────────────────────────────────────────────
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")

H1_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
H2_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
H_FONT = Font(name="微软雅黑", size=10, bold=True)

POS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NEG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

MODEL_FILLS = {
    "豆包": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "GPT": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Gemini": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
}
RESULT_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style(cell, font=None, fill=None, align=CENTER):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    cell.border = BORDER


def add_dv(ws, col, r1, r2, options):
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}{r1}:{col}{r2}")


# ── 各维度检查点定义（陈述句描述）──────────────────────────

CHECKS_SENTENCE = [
    "是否为绝对简单句（一个主语一个谓语，无任何从句）",
    "除目标词外，所有词汇均为A1基础高频词",
    "场景贴近12-15岁青少年日常（校园、家庭、天气等）",
    "若目标词有必考固定搭配，例句中已使用",
    "例句中目标词用法对应给定的具体义项和词性",
    "中文翻译与英文语义准确对应，无增减或扭曲",
    "表达自然地道，无中式英语",
    "内容阳光积极，适合K-12学生",
]

CHECKS_CHUNK = [
    "是否为高频固定搭配（动宾/形名/动介等），非随意组合",
    "目标词在语块中的用法对应给定义项和词性",
    "中文翻译自然准确，2-6个汉字",
    "表达自然地道，无中式英语逻辑",
    "是短语/词组，非完整句子",
    "内容阳光积极，适合K-12学生",
]

CHECKS_SYLLABLE = [
    "单音节词未切分 / 多音节词已正确切分",
    "每个切分块恰好包含一个发音元音（-le例外）",
    "原子单位未被拆分（sh/ch/th/ea/ee/ar/or等）",
    "静音字母（词尾silent-e、silent gh）未独立成音节",
    "辅音切分位置正确（VCV单辅音归后，VCCV从中切，双写切开）",
    "常见前缀保持完整（re-/un-/dis-/pre-/con-/ab-等）",
    "仅使用中圆点·(U+00B7)作为分隔符",
]

CHECKS_MORPHEME = [
    "词根词缀拆解有据可查（非谐音/民间词源）",
    "公式格式正确：component(中文含义) + component(中文含义)",
    "后缀标注的词性与给定词性一致",
    "口诀逻辑对应公式内容，无额外引入",
    "老师话术包含完整6步（纠音/拆解/画面/合成/裂变/升华）",
    "画面还原由词根词缀含义逻辑推出，非强行联想",
    "批量裂变词与目标词共享同一词根",
    "话术未提及任何实体教学道具（卡片/黑板/教鞭等）",
    "语气口语自然，像直播互动，非AI朗读感",
]

CHECKS_WORD_IN_WORD = [
    "内部熟词均为连续字母序列且为高频常见词",
    "所有内部词拼合后完全覆盖原词（无缺漏无多余）",
    "公式格式正确：[熟词](中文) + [熟词](中文)",
    "口诀逻辑对应公式内容，无额外引入",
    "老师话术包含完整6步（纠音/搜索/画面/合成/关联/升华）",
    "画面还原基于内部熟词构建，与目标词义有逻辑桥梁",
    "整体助记逻辑指向给定义项和词性",
    "话术未提及任何实体教学道具",
    "语气口语自然，像直播互动",
]

CHECKS_PHONETIC = [
    "谐音与英文发音高度相似（≥60%音似度）",
    "谐音为常见中文词/象声词，无消极/低俗含义",
    "公式格式正确：/IPA发音/ ≈ 中文谐音",
    "口诀逻辑对应公式内容，无额外引入",
    "老师话术包含完整6步（纠音/谐音/画面/合成/发音保护/升华）",
    "画面还原基于中文谐音构建，与词义有情感桥梁",
    "第5步明确提醒学生谐音仅为记忆辅助，不可替代正确发音",
    "整体助记逻辑指向给定义项和词性",
    "话术未提及任何实体教学道具",
    "语气口语自然，像直播互动",
]

CHECKS_EXAM = [
    "考点为真实高频中高考搭配（非泛泛而谈）",
    "公式格式正确：目标词(词性) + 核心搭配 = 中文含义(高频考点)",
    "考点逻辑精准具体，指出了搭配的核心语法规则",
    "老师话术包含完整3步（锁定考点/实战例句/避坑警示）",
    "实战例句语法正确且清晰示范了该考试搭配",
    "实战例句的中文翻译准确自然",
    "避坑警示指出了具体的错误选项（考场陷阱）",
    "话术未提及任何实体教学道具",
    "语气犀利有提分感，像考前冲刺讲师",
]

DIMENSION_CHECKS = {
    "例句": CHECKS_SENTENCE,
    "语块": CHECKS_CHUNK,
    "音节": CHECKS_SYLLABLE,
    "助记-词根词缀": CHECKS_MORPHEME,
    "助记-词中词": CHECKS_WORD_IN_WORD,
    "助记-音义联想": CHECKS_PHONETIC,
    "助记-考试应用": CHECKS_EXAM,
}

# ── 各维度的输入列定义 ────────────────────────────────────
INPUT_COLS = {
    "例句": [
        ("英文例句", 40), ("中文翻译", 25),
    ],
    "语块": [
        ("语块(英文)", 20), ("语块(中文)", 15),
    ],
    "音节": [
        ("参考正确拆分", 15), ("待测音节拆分", 18),
    ],
    "助记-词根词缀": [
        ("助记完整内容\n（核心公式+助记口诀+老师话术）", 55),
    ],
    "助记-词中词": [
        ("助记完整内容\n（核心公式+助记口诀+老师话术）", 55),
    ],
    "助记-音义联想": [
        ("助记完整内容\n（核心公式+助记口诀+老师话术）", 55),
    ],
    "助记-考试应用": [
        ("助记完整内容\n（核心公式+考点逻辑+老师话术）", 55),
    ],
}

# ── 示例数据 ──────────────────────────────────────────────
SENTENCE_SAMPLES = [
    (1, "focus (v.) 集中；聚焦", "1.常规词", "正例", "",
     "I need to focus on my homework.", "我需要专注于我的作业。",
     ["Yes"] * 8),
    (1, "focus (v.) 集中；聚焦", "1.常规词", "反例",
     "包含although从句，非绝对简单句",
     "Although I tried to focus on my homework, I couldn't stop thinking about the game.",
     "尽管我试着集中精力做作业，但我忍不住想着那场比赛。",
     ["No", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes"]),
    (21, "fine (adj.) 好的；晴朗的", "3.多义/生义", "正例", "",
     "The weather is fine today.", "今天天气很好。",
     ["Yes"] * 8),
    (21, "fine (adj.) 好的；晴朗的", "3.多义/生义", "反例",
     "例句体现\"罚款\"义，非给定的\"好的\"义项",
     "He was fined 200 dollars for parking.", "他因停车被罚了200美元。",
     ["Yes", "Yes", "Yes", "Yes", "No", "Yes", "Yes", "Yes"]),
    (34, "affect (v.) 影响", "4.易混淆", "正例", "",
     "The rain will affect our plan.", "雨会影响我们的计划。",
     ["Yes"] * 8),
    (34, "affect (v.) 影响", "4.易混淆", "反例",
     "affect误用为名词(应为effect)，且significantly超出A1",
     "The affect of the rain was significant.", "雨的影响是显著的。",
     ["Yes", "No", "Yes", "Yes", "No", "Yes", "No", "Yes"]),
    (45, "abstract (adj.) 抽象的", "5.高阶抽象", "正例", "",
     "Art can be abstract.", "艺术可以是抽象的。",
     ["Yes"] * 8),
    (45, "abstract (adj.) 抽象的", "5.高阶抽象", "反例",
     "ramifications/negligence远超A1基础词汇",
     "The ramifications of negligence led to abstract consequences.",
     "疏忽的后果导致了抽象的后果。",
     ["Yes", "No", "Yes", "Yes", "Yes", "Yes", "No", "Yes"]),
]

CHUNK_SAMPLES = [
    (1, "focus (v.) 集中；聚焦", "1.常规词", "正例", "",
     "focus on", "专注于",
     ["Yes"] * 6),
    (1, "focus (v.) 集中；聚焦", "1.常规词", "反例",
     "focus a thing不是固定搭配，应为focus on",
     "focus a thing", "集中一个东西",
     ["No", "Yes", "Yes", "No", "Yes", "Yes"]),
    (21, "fine (adj.) 好的；晴朗的", "3.多义/生义", "正例", "",
     "fine weather", "好天气",
     ["Yes"] * 6),
    (21, "fine (adj.) 好的；晴朗的", "3.多义/生义", "反例",
     "pay a fine对应\"罚款\"义，非给定的\"好的\"义项",
     "pay a fine", "付罚款",
     ["Yes", "No", "Yes", "Yes", "Yes", "Yes"]),
]

SYLLABLE_SAMPLES = [
    (1, "focus", "1.常规词", "正例", "",
     "fo·cus", "fo·cus",
     ["Yes"] * 7),
    (1, "focus", "1.常规词", "反例",
     "VCV模式单辅音c应归后(fo·cus)，而非归前(foc·us)",
     "fo·cus", "foc·us",
     ["Yes", "Yes", "Yes", "Yes", "No", "Yes", "Yes"]),
    (7, "protect", "1.常规词", "正例", "",
     "pro·tect", "pro·tect",
     ["Yes"] * 7),
    (7, "protect", "1.常规词", "反例",
     "前缀pro-被破坏，拆成了prot·ect",
     "pro·tect", "prot·ect",
     ["Yes", "Yes", "Yes", "Yes", "Yes", "No", "Yes"]),
    (15, "deep", "2.不规则变形", "正例", "",
     "deep", "deep",
     ["Yes"] * 7),
    (15, "deep", "2.不规则变形", "反例",
     "单音节词不应切分，且ee元音团被拆开",
     "deep", "de·ep",
     ["No", "No", "No", "Yes", "Yes", "Yes", "Yes"]),
    (31, "consistent", "4.易混淆", "正例", "",
     "con·sis·tent", "con·sis·tent",
     ["Yes"] * 7),
    (31, "consistent", "4.易混淆", "反例",
     "前缀con-被破坏，拆成了cons·is·tent",
     "con·sis·tent", "cons·is·tent",
     ["Yes", "Yes", "Yes", "Yes", "Yes", "No", "Yes"]),
    (45, "abstract", "5.高阶抽象", "正例", "",
     "ab·stract", "ab·stract",
     ["Yes"] * 7),
    (45, "abstract", "5.高阶抽象", "反例",
     "前缀ab-被破坏，拆成了abs·tract",
     "ab·stract", "abs·tract",
     ["Yes", "Yes", "Yes", "Yes", "Yes", "No", "Yes"]),
]

SAMPLES = {
    "例句": SENTENCE_SAMPLES,
    "语块": CHUNK_SAMPLES,
    "音节": SYLLABLE_SAMPLES,
    "助记-词根词缀": [],
    "助记-词中词": [],
    "助记-音义联想": [],
    "助记-考试应用": [],
}


# ═══════════════════════════════════════════════════════════
#  案例表
# ═══════════════════════════════════════════════════════════

def build_case_sheet(wb, dim_name):
    """为一个维度创建案例 sheet."""
    ws = wb.create_sheet(dim_name)
    checks = DIMENSION_CHECKS[dim_name]
    inputs = INPUT_COLS[dim_name]
    samples = SAMPLES.get(dim_name, [])

    # ── Row 1: 大标题
    total_cols = 6 + len(inputs) + len(checks) + 1  # 基础6 + 输入 + 检查点 + 期望整体
    end_col = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value = f"质检黄金案例表 · {dim_name}维度"
    style(c, TITLE_FONT, TITLE_FILL)
    ws.row_dimensions[1].height = 35

    # ── Row 2: 分区标题
    base_end = get_column_letter(5 + len(inputs))
    ws.merge_cells(f"A2:{base_end}2")
    style(ws["A2"], H_FONT, H1_FILL)
    ws["A2"].value = "基础信息 + 输入内容"

    check_start = get_column_letter(6 + len(inputs))
    check_end = get_column_letter(5 + len(inputs) + len(checks))
    overall_col = get_column_letter(6 + len(inputs) + len(checks))
    ws.merge_cells(f"{check_start}2:{overall_col}2")
    style(ws[f"{check_start}2"], H_FONT, RESULT_FILL)
    ws[f"{check_start}2"].value = "各检查点期望结果（人工标注 Ground Truth）"

    # ── Row 3: 列标题
    base_headers = [
        ("A", "序号", 6),
        ("B", "测试词\n单词/词性/释义", 20),
        ("C", "类别", 10),
        ("D", "样本类型\n（正例/反例）", 10),
        ("E", "反例问题描述\n（正例留空）", 30),
    ]

    col_idx = 6
    for name, width in inputs:
        letter = get_column_letter(col_idx)
        base_headers.append((letter, name, width))
        col_idx += 1

    for ch in checks:
        letter = get_column_letter(col_idx)
        base_headers.append((letter, ch, 18))
        col_idx += 1

    base_headers.append((get_column_letter(col_idx), "期望整体结果\n（PASS/FAIL）", 12))

    for letter, header, width in base_headers:
        cell = ws[f"{letter}3"]
        cell.value = header
        style(cell, H_FONT, H2_FILL, LEFT if len(header) > 20 else CENTER)
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[3].height = 65

    # ── Row 4+: 示例数据
    row = 4
    input_count = len(inputs)
    check_count = len(checks)

    for sample in samples:
        seq, word, cat, stype, desc = sample[:5]
        input_vals = sample[5:5 + input_count]
        check_vals = sample[5 + input_count]

        ws[f"A{row}"] = seq
        ws[f"B{row}"] = word
        ws[f"C{row}"] = cat
        ws[f"D{row}"] = stype
        ws[f"E{row}"] = desc

        ci = 6
        for val in input_vals:
            style(ws[f"{get_column_letter(ci)}{row}"],
                  fill=POS_FILL if stype == "正例" else NEG_FILL, align=LEFT)
            ws[f"{get_column_letter(ci)}{row}"] = val
            ci += 1

        for i, val in enumerate(check_vals):
            letter = get_column_letter(ci + i)
            ws[f"{letter}{row}"] = val
            fill = POS_FILL if val == "Yes" else NEG_FILL if val == "No" else None
            style(ws[f"{letter}{row}"], fill=fill)

        overall_letter = get_column_letter(ci + check_count)
        overall = "PASS" if all(v == "Yes" for v in check_vals) else "FAIL"
        ws[f"{overall_letter}{row}"] = overall
        style(ws[f"{overall_letter}{row}"],
              fill=POS_FILL if overall == "PASS" else NEG_FILL,
              font=Font(bold=True, name="微软雅黑"))

        # 基础列样式
        bg = POS_FILL if stype == "正例" else NEG_FILL
        for letter in ["A", "B", "C", "D", "E"]:
            style(ws[f"{letter}{row}"], fill=bg,
                  align=LEFT if letter in "BE" else CENTER)

        ws.row_dimensions[row].height = 55 if dim_name.startswith("助记") else 40
        row += 1

    # 占位提示
    if not samples:
        ws[f"A{row}"] = 1
        ws[f"B{row}"] = "（示例词）"
        ws[f"D{row}"] = "正例"
        ws[f"E{row}"] = ""
        ci = 6
        for _ in inputs:
            ws[f"{get_column_letter(ci)}{row}"] = "（从生产评估表中提取最优模型输出）"
            ci += 1
        row += 1
        ws[f"A{row}"] = 1
        ws[f"B{row}"] = "（示例词）"
        ws[f"D{row}"] = "反例"
        ws[f"E{row}"] = "（用陈述句描述：在正例基础上改了什么、为什么应该不通过）"
        ci = 6
        for _ in inputs:
            ws[f"{get_column_letter(ci)}{row}"] = "（在正例基础上做定向篡改后的内容）"
            ci += 1

    # 数据验证
    add_dv(ws, "D", 4, 300, ["正例", "反例"])
    overall_col_letter = get_column_letter(6 + input_count + check_count)
    add_dv(ws, overall_col_letter, 4, 300, ["PASS", "FAIL"])
    for i in range(check_count):
        letter = get_column_letter(6 + input_count + i)
        add_dv(ws, letter, 4, 300, ["Yes", "No", "NA"])

    ws.freeze_panes = "A4"


def build_case_book():
    wb = Workbook()
    wb.remove(wb.active)

    # 封面/说明 sheet
    ws = wb.create_sheet("说明")
    ws.merge_cells("A1:F1")
    style(ws["A1"], TITLE_FONT, TITLE_FILL)
    ws["A1"].value = "质检黄金案例表 — 使用说明"
    ws.row_dimensions[1].height = 35
    ws.column_dimensions["A"].width = 60

    notes = [
        "本表定义所有维度的正例+反例测试数据，作为质检模型评估的 Ground Truth。",
        "",
        "■ 每个维度一个 Sheet，每个词2行（正例+反例）",
        "",
        "■ 正例来源：从已完成的「生产模型评估表」中提取最优模型的输出",
        "  → 例句取 sentence sheet 最终选择的模型输出",
        "  → 语块取 chunk sheet 最终选择的模型输出",
        "  → 音节取词表 F 列已有的正确拆分",
        "  → 助记取各助记 sheet 最终选择的模型输出",
        "",
        "■ 反例构造：在正例基础上做【最小化定向篡改】",
        "  → 一次只改一个点，让差异最小化",
        "  → 「反例问题描述」列用陈述句说明改了什么",
        "  → 对应的检查点列标为 No",
        "",
        "■ 各检查点列的值：",
        "  Yes = 该检查点通过",
        "  No  = 该检查点不通过（反例的定向违反点）",
        "  NA  = 该检查点不适用",
        "",
        "■ 「期望整体结果」：所有检查点均 Yes → PASS，任一 No → FAIL",
        "",
        "■ 50 词 × 5 类别：",
        "  1.常规词(10) | 2.不规则变形(10) | 3.多义/生义(10) | 4.易混淆(10) | 5.高阶抽象(10)",
    ]
    for i, line in enumerate(notes):
        ws[f"A{3 + i}"] = line

    for dim_name in DIMENSION_CHECKS:
        build_case_sheet(wb, dim_name)

    return wb


# ═══════════════════════════════════════════════════════════
#  评估表
# ═══════════════════════════════════════════════════════════

MODELS = ["豆包", "GPT", "Gemini"]


def build_eval_sheet(wb, dim_name):
    """为一个维度创建评估 sheet.

    每个词的正例/反例各展开 3 行（豆包/GPT/Gemini），与生产评估表排列一致：
    focus 正例  豆包   Yes Yes ... PASS ✅
    focus 正例  GPT    Yes Yes ... PASS ✅
    focus 正例  Gemini Yes No  ... FAIL ❌
    focus 反例  豆包   Yes No  ... FAIL ✅
    focus 反例  GPT    Yes Yes ... PASS ❌ ← 漏判
    focus 反例  Gemini Yes No  ... FAIL ✅
    """
    ws = wb.create_sheet(dim_name)
    checks = DIMENSION_CHECKS[dim_name]
    check_count = len(checks)

    # 列结构: 基础(7列) + 各检查点(N列) + 整体判断 + 是否正确 + 模型输出理由 + 备注
    total_cols = 7 + check_count + 4
    end_col = get_column_letter(total_cols)

    # Row 1: 大标题
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value = f"质检模型评估表 · {dim_name}维度"
    style(c, TITLE_FONT, TITLE_FILL)
    ws.row_dimensions[1].height = 35

    # Row 2: 分区标题
    ws.merge_cells("A2:G2")
    style(ws["A2"], H_FONT, H1_FILL)
    ws["A2"].value = "案例信息"

    check_start = get_column_letter(8)
    check_end = get_column_letter(7 + check_count)
    ws.merge_cells(f"{check_start}2:{check_end}2")
    style(ws[f"{check_start}2"], H_FONT, RESULT_FILL)
    ws[f"{check_start}2"].value = "模型对各检查点的判断（Yes=通过  No=不通过  NA=不适用）"

    result_start = get_column_letter(8 + check_count)
    result_end = end_col
    ws.merge_cells(f"{result_start}2:{result_end}2")
    style(ws[f"{result_start}2"], H_FONT, SUMMARY_FILL)
    ws[f"{result_start}2"].value = "汇总与结论"

    # Row 3: 列标题
    base_headers = [
        ("A", "序号", 6),
        ("B", "测试词\n单词/词性/释义", 18),
        ("C", "类别", 10),
        ("D", "样本类型\n正例/反例", 9),
        ("E", "反例问题描述", 28),
        ("F", "期望整体\nPASS/FAIL", 9),
        ("G", "测试模型", 9),
    ]

    for letter, header, width in base_headers:
        cell = ws[f"{letter}3"]
        cell.value = header
        style(cell, H_FONT, H2_FILL, LEFT if letter == "E" else CENTER)
        ws.column_dimensions[letter].width = width

    col_idx = 8
    for ch in checks:
        letter = get_column_letter(col_idx)
        short = ch[:15] + "…" if len(ch) > 17 else ch
        ws[f"{letter}3"] = short
        style(ws[f"{letter}3"], H_FONT, H2_FILL, CENTER)
        ws.column_dimensions[letter].width = 14
        col_idx += 1

    tail_headers = [
        ("整体判断\nPASS/FAIL", 10),
        ("是否正确\n✅/❌", 8),
        ("模型输出理由", 28),
        ("备注", 18),
    ]
    for header, width in tail_headers:
        letter = get_column_letter(col_idx)
        ws[f"{letter}3"] = header
        sfill = SUMMARY_FILL
        style(ws[f"{letter}3"], H_FONT, sfill, CENTER)
        ws.column_dimensions[letter].width = width
        col_idx += 1

    ws.row_dimensions[3].height = 65

    # Row 4: 检查点全称提示行
    col_idx = 8
    for ch in checks:
        letter = get_column_letter(col_idx)
        ws[f"{letter}4"] = ch
        style(ws[f"{letter}4"], Font(name="微软雅黑", size=8, color="666666"), align=LEFT)
        col_idx += 1
    ws.row_dimensions[4].height = 75

    # Row 5+: 预填示例数据（例句维度展示完整示例）
    data_start_row = 5
    if dim_name == "例句":
        samples = [
            # (seq, word, cat, stype, desc, expected)
            (1, "focus (v.)\n集中；聚焦", "1.常规词", "正例", "", "PASS"),
            (21, "fine (adj.)\n好的；晴朗的", "3.多义/生义", "正例", "", "PASS"),
            (21, "fine (adj.)\n好的；晴朗的", "3.多义/生义", "反例",
             "例句体现\"罚款\"义，非给定的\"好的\"义项", "FAIL"),
        ]
        row = data_start_row
        for seq, word, cat, stype, desc, expected in samples:
            for model in MODELS:
                ws[f"A{row}"] = seq
                ws[f"B{row}"] = word
                ws[f"C{row}"] = cat
                ws[f"D{row}"] = stype
                ws[f"E{row}"] = desc
                ws[f"F{row}"] = expected
                ws[f"G{row}"] = model

                # 行背景色：正例绿底、反例橙底，模型列用模型色
                bg = POS_FILL if stype == "正例" else NEG_FILL
                for letter in "ABCDEF":
                    style(ws[f"{letter}{row}"], fill=bg,
                          align=LEFT if letter in "BE" else CENTER)
                style(ws[f"G{row}"], fill=MODEL_FILLS[model])

                # 检查点列和汇总列留空待填
                for ci in range(8, 8 + check_count + 4):
                    style(ws[f"{get_column_letter(ci)}{row}"])

                ws.row_dimensions[row].height = 35
                row += 1

        # 占位提示
        ws[f"A{row}"] = "..."
        ws[f"B{row}"] = "（50词 × 正反例 × 3模型 = 300行）"
        ws.merge_cells(f"B{row}:G{row}")

    # 数据验证
    max_row = 600
    add_dv(ws, "D", data_start_row, max_row, ["正例", "反例"])
    add_dv(ws, "F", data_start_row, max_row, ["PASS", "FAIL"])
    add_dv(ws, "G", data_start_row, max_row, MODELS)

    for i in range(check_count):
        letter = get_column_letter(8 + i)
        add_dv(ws, letter, data_start_row, max_row, ["Yes", "No", "NA"])

    overall_col = get_column_letter(8 + check_count)
    correct_col = get_column_letter(9 + check_count)
    add_dv(ws, overall_col, data_start_row, max_row, ["PASS", "FAIL"])
    add_dv(ws, correct_col, data_start_row, max_row, ["✅", "❌"])

    ws.freeze_panes = "H5"


def build_summary_sheet(wb):
    ws = wb.create_sheet("汇总")

    ws.merge_cells("A1:I1")
    style(ws["A1"], TITLE_FONT, TITLE_FILL)
    ws["A1"].value = "质检模型选择 · 汇总报告"
    ws.row_dimensions[1].height = 35
    ws.column_dimensions["A"].width = 14

    # ── 表1
    ws["A3"] = "表1: 模型 × 维度 准确率"
    style(ws["A3"], Font(name="微软雅黑", size=12, bold=True), align=LEFT)

    dims = ["例句", "语块", "音节", "词根词缀", "词中词", "音义联想", "考试应用", "平均"]
    r = 4
    ws[f"A{r}"] = "模型"
    style(ws[f"A{r}"], H_FONT, H2_FILL)
    for i, d in enumerate(dims):
        col = get_column_letter(i + 2)
        ws[f"{col}{r}"] = d
        style(ws[f"{col}{r}"], H_FONT, H2_FILL)
        ws.column_dimensions[col].width = 10

    for ri, m in enumerate(MODELS, start=5):
        ws[f"A{ri}"] = m
        style(ws[f"A{ri}"], H_FONT, MODEL_FILLS[m])
        for i in range(len(dims)):
            style(ws[f"{get_column_letter(i + 2)}{ri}"])

    # ── 表2
    ws["A9"] = "表2: 模型 × 词类别 准确率（多义词和易混淆词最能区分模型能力）"
    style(ws["A9"], Font(name="微软雅黑", size=12, bold=True), align=LEFT)

    cats = ["常规词", "不规则变形", "多义/生义 ⭐", "易混淆 ⭐", "高阶抽象", "平均"]
    r = 10
    ws[f"A{r}"] = "模型"
    style(ws[f"A{r}"], H_FONT, H2_FILL)
    for i, c in enumerate(cats):
        col = get_column_letter(i + 2)
        ws[f"{col}{r}"] = c
        style(ws[f"{col}{r}"], H_FONT, H2_FILL)
        ws.column_dimensions[col].width = 13

    for ri, m in enumerate(MODELS, start=11):
        ws[f"A{ri}"] = m
        style(ws[f"A{ri}"], H_FONT, MODEL_FILLS[m])
        for i in range(len(cats)):
            style(ws[f"{get_column_letter(i + 2)}{ri}"])

    # ── 表3
    ws["A15"] = "表3: 假阳率 FPR = 反例中被判为PASS的比例（越低越好，0%最理想）"
    style(ws["A15"], Font(name="微软雅黑", size=12, bold=True, color="CC0000"), align=LEFT)

    r = 16
    ws[f"A{r}"] = "模型"
    style(ws[f"A{r}"], H_FONT, H2_FILL)
    fpr_dims = ["例句", "语块", "音节", "词根词缀", "词中词", "音义联想", "考试应用", "总体"]
    for i, d in enumerate(fpr_dims):
        col = get_column_letter(i + 2)
        ws[f"{col}{r}"] = d
        style(ws[f"{col}{r}"], H_FONT, RESULT_FILL)

    for ri, m in enumerate(MODELS, start=17):
        ws[f"A{ri}"] = m
        style(ws[f"A{ri}"], H_FONT, MODEL_FILLS[m])
        for i in range(len(fpr_dims)):
            style(ws[f"{get_column_letter(i + 2)}{ri}"])

    # ── 表4
    ws["A21"] = "表4: 假阴率 FNR = 正例中被判为FAIL的比例（误杀率）"
    style(ws["A21"], Font(name="微软雅黑", size=12, bold=True), align=LEFT)

    r = 22
    ws[f"A{r}"] = "模型"
    style(ws[f"A{r}"], H_FONT, H2_FILL)
    for i, d in enumerate(fpr_dims):
        col = get_column_letter(i + 2)
        ws[f"{col}{r}"] = d
        style(ws[f"{col}{r}"], H_FONT, H2_FILL)

    for ri, m in enumerate(MODELS, start=23):
        ws[f"A{ri}"] = m
        style(ws[f"A{ri}"], H_FONT, MODEL_FILLS[m])
        for i in range(len(fpr_dims)):
            style(ws[f"{get_column_letter(i + 2)}{ri}"])

    # ── 表5: 最终推荐
    ws["A27"] = "表5: 最终推荐（系统支持按维度配置不同模型）"
    style(ws["A27"], Font(name="微软雅黑", size=12, bold=True), align=LEFT)

    rec = ["维度", "推荐模型", "准确率", "假阳率", "假阴率", "备选模型", "选择理由"]
    for i, h in enumerate(rec):
        col = get_column_letter(i + 1)
        ws[f"{col}28"] = h
        style(ws[f"{col}28"], H_FONT, SUMMARY_FILL)
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width, 12)

    for ri, d in enumerate(["例句", "语块", "音节", "助记-词根词缀",
                            "助记-词中词", "助记-音义联想", "助记-考试应用"], start=29):
        ws[f"A{ri}"] = d
        style(ws[f"A{ri}"], H_FONT)
        for i in range(1, len(rec)):
            style(ws[f"{get_column_letter(i + 1)}{ri}"])

    # ── 说明
    ws["A38"] = "指标说明"
    style(ws["A38"], Font(name="微软雅黑", size=12, bold=True), align=LEFT)
    notes = [
        "准确率 = (正例正确判PASS数 + 反例正确判FAIL数) ÷ 总样本数",
        "假阳率 FPR = 反例被模型判为PASS的数量 ÷ 反例总数  ← 最关键，越低越好",
        "假阴率 FNR = 正例被模型判为FAIL的数量 ÷ 正例总数  ← 误杀率",
        "产品红线：FPR必须低，宁可误杀（FNR高），不可放行坏数据",
        "多义/生义词 + 易混淆词的准确率最能区分模型能力",
    ]
    for i, n in enumerate(notes):
        ws[f"A{39 + i}"] = n

    ws.freeze_panes = "A4"


def build_eval_book():
    wb = Workbook()
    wb.remove(wb.active)

    # 说明 sheet
    ws = wb.create_sheet("说明")
    ws.merge_cells("A1:F1")
    style(ws["A1"], TITLE_FONT, TITLE_FILL)
    ws["A1"].value = "质检模型评估表 — 使用说明"
    ws.row_dimensions[1].height = 35
    ws.column_dimensions["A"].width = 65

    notes = [
        "本表记录3个候选模型对黄金案例的判断结果，与 Ground Truth 对比。",
        "",
        "■ 数据排列方式（与生产评估表一致）：",
        "  每个词的正例/反例各展开 3 行（豆包/GPT/Gemini）：",
        "  focus  正例  豆包   Yes Yes Yes ... PASS  ✅",
        "  focus  正例  GPT    Yes Yes Yes ... PASS  ✅",
        "  focus  正例  Gemini Yes Yes No  ... FAIL  ❌ ← 误杀",
        "  focus  反例  豆包   Yes Yes No  ... FAIL  ✅",
        "  focus  反例  GPT    Yes Yes Yes ... PASS  ❌ ← 漏判（假阳）",
        "  focus  反例  Gemini Yes Yes No  ... FAIL  ✅",
        "",
        "  50词 × 2(正反例) × 3(模型) = 300 行/维度",
        "",
        "■ 使用流程：",
        "  1. 先完成「质检黄金案例表」（正例+反例 + 各检查点期望结果）",
        "  2. 将同一份案例分别喂给 3 个模型，记录各检查点的 Yes/No 判断",
        "  3. 对比模型整体判断 vs 案例表的期望结果 → 填写「是否正确」列",
        "  4. 在汇总 sheet 统计各指标",
        "",
        "■ Row 4 是检查点全称提示行（可隐藏），Row 3 是缩略列头",
        "",
        "■ 「是否正确」判断标准：",
        "  模型整体判断 == 期望整体结果 → ✅",
        "  模型整体判断 != 期望整体结果 → ❌",
        "  特别关注：反例被判PASS = ❌ = 假阳 = 坏数据被放行（最严重的错误）",
    ]
    for i, line in enumerate(notes):
        ws[f"A{3 + i}"] = line

    for dim_name in DIMENSION_CHECKS:
        build_eval_sheet(wb, dim_name)

    build_summary_sheet(wb)
    return wb


# ── 主函数 ────────────────────────────────────────────────
def main():
    case_wb = build_case_book()
    case_path = "docs/design/质检黄金案例表.xlsx"
    case_wb.save(case_path)

    eval_wb = build_eval_book()
    eval_path = "docs/design/质检模型评估表.xlsx"
    eval_wb.save(eval_path)

    print(f"✅ 案例表: {case_path}")
    print(f"   Sheets: {', '.join(case_wb.sheetnames)}")
    print(f"✅ 评估表: {eval_path}")
    print(f"   Sheets: {', '.join(eval_wb.sheetnames)}")


if __name__ == "__main__":
    main()
