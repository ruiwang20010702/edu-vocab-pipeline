"""从黄金案例表 xlsx 提取测试案例为结构化数据。"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl

from config import DIMENSION_SHEET_NAMES, GOLDEN_XLSX, RESULTS_DIR


@dataclass
class TestCase:
    dimension: str          # sentence / chunk / syllable / mnemonic_*
    case_id: int            # 序号
    word: str               # 测试词 (e.g. "focus (v.) 集中；聚焦")
    category: str           # 类别 (e.g. "常规词")
    sample_type: str        # "正例" / "反例"
    issue_desc: str         # 反例问题描述
    expected: str           # "PASS" / "FAIL"
    user_prompt_input: str  # 拼好的 user prompt 输入


def _parse_word_info(word_str: str) -> tuple[str, str, str]:
    """从 'focus (v.) 集中；聚焦' 解析出 word, pos, meaning。"""
    word_str = word_str.strip()
    # Try to split: "focus (v.) 集中；聚焦" or "focus  (v.) 集中；聚焦"
    import re
    m = re.match(r"^(\S+)\s+\(([^)]+)\)\s*(.*)", word_str)
    if m:
        return m.group(1), m.group(2), m.group(3)
    # Fallback: just word
    parts = word_str.split()
    return parts[0] if parts else word_str, "", ""


def _extract_sentence(ws) -> list[TestCase]:
    """提取例句维度案例。"""
    cases = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        seq = row[0]
        if seq is None or str(seq).strip() in ("", "..."):
            continue
        word_raw = str(row[1] or "")
        category = str(row[2] or "")
        sample_type = str(row[3] or "")
        issue_desc = str(row[4] or "")
        en_sentence = str(row[5] or "")
        cn_translation = str(row[6] or "")

        word, pos, meaning = _parse_word_info(word_raw)

        # 找期望结果列 (PASS/FAIL)
        expected = str(row[15] or "").strip().upper()
        if expected not in ("PASS", "FAIL"):
            # 尝试其他列
            for col_idx in range(15, min(19, len(row))):
                val = str(row[col_idx] or "").strip().upper()
                if val in ("PASS", "FAIL"):
                    expected = val
                    break

        user_input = (
            f"Word: {word} | POS: {pos or '未知'} | "
            f"Meaning: {meaning or '无'} | "
            f"Sentence: {en_sentence} | Chinese: {cn_translation}"
        )

        cases.append(TestCase(
            dimension="sentence",
            case_id=int(seq),
            word=word_raw,
            category=category,
            sample_type="正例" if "正例" in sample_type else "反例",
            issue_desc=issue_desc,
            expected=expected,
            user_prompt_input=user_input,
        ))
    return cases


def _extract_chunk(ws) -> list[TestCase]:
    """提取语块维度案例。"""
    cases = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        seq = row[0]
        if seq is None or str(seq).strip() in ("", "..."):
            continue
        word_raw = str(row[1] or "")
        category = str(row[2] or "")
        sample_type = str(row[3] or "")
        issue_desc = str(row[4] or "")
        chunk_en = str(row[5] or "")
        chunk_cn = str(row[6] or "")

        word, pos, meaning = _parse_word_info(word_raw)

        expected = str(row[13] or "").strip().upper()
        if expected not in ("PASS", "FAIL"):
            for col_idx in range(13, min(16, len(row))):
                val = str(row[col_idx] or "").strip().upper()
                if val in ("PASS", "FAIL"):
                    expected = val
                    break

        user_input = (
            f"Word: {word} | POS: {pos or '未知'} | "
            f"Meaning: {meaning or '无'} | "
            f"Chunk: {chunk_en} | Chinese: {chunk_cn}"
        )

        cases.append(TestCase(
            dimension="chunk",
            case_id=int(seq),
            word=word_raw,
            category=category,
            sample_type="正例" if "正例" in sample_type else "反例",
            issue_desc=issue_desc,
            expected=expected,
            user_prompt_input=user_input,
        ))
    return cases


def _extract_syllable(ws) -> list[TestCase]:
    """提取音节维度案例。"""
    cases = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        seq = row[0]
        if seq is None or str(seq).strip() in ("", "..."):
            continue
        word_raw = str(row[1] or "")
        category = str(row[2] or "")
        sample_type = str(row[3] or "")
        issue_desc = str(row[4] or "")
        _correct_split = str(row[5] or "")
        test_split = str(row[6] or "")

        word = word_raw.strip().split()[0] if word_raw.strip() else word_raw

        expected = str(row[14] or "").strip().upper()
        if expected not in ("PASS", "FAIL"):
            for col_idx in range(14, min(17, len(row))):
                val = str(row[col_idx] or "").strip().upper()
                if val in ("PASS", "FAIL"):
                    expected = val
                    break

        user_input = f"Word: {word} | Segmentation: {test_split}"

        cases.append(TestCase(
            dimension="syllable",
            case_id=int(seq),
            word=word_raw,
            category=category,
            sample_type="正例" if "正例" in sample_type else "反例",
            issue_desc=issue_desc,
            expected=expected,
            user_prompt_input=user_input,
        ))
    return cases


def _extract_mnemonic(ws, dimension: str) -> list[TestCase]:
    """提取助记维度案例（4 种助记共用逻辑）。"""
    cases = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        seq = row[0]
        if seq is None or str(seq).strip() in ("", "..."):
            continue
        word_raw = str(row[1] or "")
        category = str(row[2] or "")
        # 助记表有 "负责人" 列在 index 3
        sample_type = str(row[4] or "")
        issue_desc = str(row[5] or "")
        content = str(row[6] or "")

        # 跳过没有内容的行（正例全 NA 的词）
        if not content or content.strip() in ("", "False", "NA", "(empty)"):
            # 检查是否为正例且无内容——该词不适用此助记类型
            if "正例" in sample_type:
                continue
            # 反例也需要有内容才能测
            if not content or content.strip() in ("", "False", "NA"):
                continue

        word, pos, meaning = _parse_word_info(word_raw)

        # 期望结果列位置因维度不同而异，搜索最后几列
        expected = ""
        for col_idx in range(len(row) - 1, max(len(row) - 5, 6), -1):
            val = str(row[col_idx] or "").strip().upper()
            if val in ("PASS", "FAIL"):
                expected = val
                break

        if not expected:
            continue

        # 构建 user prompt
        if "考试" in dimension or "exam" in dimension:
            user_input = (
                f"Word: {word} | POS: {pos or '未知'} | "
                f"Meaning: {meaning or '无'} | Mnemonic: {content}"
            )
        else:
            user_input = (
                f"Word: {word} | POS: {pos or '未知'} | "
                f"Meaning: {meaning or '无'} | Mnemonic: {content}"
            )

        cases.append(TestCase(
            dimension=dimension,
            case_id=int(seq),
            word=word_raw,
            category=category,
            sample_type="正例" if "正例" in sample_type else "反例",
            issue_desc=issue_desc,
            expected=expected,
            user_prompt_input=user_input,
        ))
    return cases


EXTRACTORS = {
    "sentence": _extract_sentence,
    "chunk": _extract_chunk,
    "syllable": _extract_syllable,
}


def extract_all_cases() -> dict[str, list[TestCase]]:
    """从黄金案例表提取所有维度的测试案例。"""
    wb = openpyxl.load_workbook(GOLDEN_XLSX, data_only=True)
    all_cases: dict[str, list[TestCase]] = {}

    for dimension, sheet_name in DIMENSION_SHEET_NAMES.items():
        ws = wb[sheet_name]
        if dimension in EXTRACTORS:
            cases = EXTRACTORS[dimension](ws)
        else:
            cases = _extract_mnemonic(ws, dimension)
        all_cases[dimension] = cases

    wb.close()
    return all_cases


def save_cases_json(all_cases: dict[str, list[TestCase]]) -> Path:
    """将提取的案例保存为 JSON 文件。"""
    RESULTS_DIR.mkdir(exist_ok=True)
    out_path = RESULTS_DIR / "test_cases.json"
    data = {
        dim: [asdict(c) for c in cases]
        for dim, cases in all_cases.items()
    }
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


if __name__ == "__main__":
    all_cases = extract_all_cases()
    for dim, cases in all_cases.items():
        pos = sum(1 for c in cases if c.sample_type == "正例")
        neg = sum(1 for c in cases if c.sample_type == "反例")
        print(f"{dim}: {len(cases)} cases ({pos} 正例 + {neg} 反例)")

    out = save_cases_json(all_cases)
    print(f"\nSaved to {out}")
