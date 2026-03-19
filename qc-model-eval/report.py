"""汇总评估结果，生成统计报告 + xlsx 评估表。

xlsx 格式与「质检模型评估表」对齐：
- 每维度一个 sheet
- 每个案例 × 3 模型 = 3 行
- 列：案例信息 | 检查点 Yes/No | 整体判断 | 是否正确 | 模型输出理由 | 备注
- 汇总 sheet：准确率/FPR/FNR/推荐
"""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from checkpoints import DIMENSION_CHECKPOINTS
from config import RESULTS_DIR

# ── 样式 ──────────────────────────────────────────────
TITLE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YES_FONT = Font(color="006100")
NO_FONT = Font(color="9C0006", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

DIMENSION_CN = {
    "sentence": "例句",
    "chunk": "语块",
    "syllable": "音节",
    "mnemonic_root_affix": "词根词缀",
    "mnemonic_word_in_word": "词中词",
    "mnemonic_sound_meaning": "音义联想",
    "mnemonic_exam_app": "考试应用",
}

MODEL_ORDER = ["GPT-5.2", "Gemini", "豆包"]


def load_all_results() -> dict[str, dict[str, list[dict]]]:
    """加载所有评估结果。返回 {dimension: {model: [results]}}。"""
    all_data: dict[str, dict[str, list[dict]]] = defaultdict(dict)
    for path in sorted(RESULTS_DIR.glob("eval_*.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        dim = data["dimension"]
        model = data["model"]
        all_data[dim][model] = data["results"]
    return dict(all_data)


def compute_metrics(results: list[dict]) -> dict:
    total = len(results)
    if total == 0:
        return {"accuracy": 0, "fpr": 0, "fnr": 0, "total": 0}
    correct = sum(1 for r in results if r.get("is_correct"))
    positives = [r for r in results if r["sample_type"] == "正例"]
    negatives = [r for r in results if r["sample_type"] == "反例"]
    false_positives = sum(1 for r in negatives if r.get("model_judgment") == "PASS")
    false_negatives = sum(1 for r in positives if r.get("model_judgment") == "FAIL")
    fpr = false_positives / len(negatives) if negatives else 0
    fnr = false_negatives / len(positives) if positives else 0
    return {
        "accuracy": correct / total, "fpr": fpr, "fnr": fnr,
        "total": total, "correct": correct,
        "positives": len(positives), "negatives": len(negatives),
        "false_positives": false_positives, "false_negatives": false_negatives,
    }


# ── 每维度 detail sheet ──────────────────────────────────
def _write_dimension_sheet(
    wb: openpyxl.Workbook,
    dimension: str,
    dim_data: dict[str, list[dict]],
    models: list[str],
) -> None:
    dim_cn = DIMENSION_CN.get(dimension, dimension)
    checkpoints = DIMENSION_CHECKPOINTS.get(dimension, [])
    ws = wb.create_sheet(title=dim_cn)

    # ── Row 1: 标题 ──
    title = f"质检模型评估表 · {dim_cn}维度"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7 + len(checkpoints) + 3)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # ── Row 2: 分区标题 ──
    info_end = 7  # A-G: 案例信息
    cp_start = 8
    cp_end = cp_start + len(checkpoints) - 1
    summary_start = cp_end + 1

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=info_end)
    c = ws.cell(row=2, column=1, value="案例信息")
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = CENTER

    ws.merge_cells(start_row=2, start_column=cp_start, end_row=2, end_column=cp_end)
    c = ws.cell(row=2, column=cp_start, value="模型对各检查点的判断（Yes=通过  No=不通过  NA=不适用）")
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = CENTER

    ws.merge_cells(start_row=2, start_column=summary_start, end_row=2, end_column=summary_start + 2)
    c = ws.cell(row=2, column=summary_start, value="汇总与结论")
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = CENTER

    # ── Row 3: 列头 ──
    col_headers = [
        "序号", "测试词\n单词/词性/释义", "类别",
        "样本类型\n正例/反例", "反例问题描述",
        "期望整体\nPASS/FAIL", "测试模型",
    ]
    for cp in checkpoints:
        # 缩短列头：取前 8 个字
        short = cp.name_cn[:8] + "…" if len(cp.name_cn) > 8 else cp.name_cn
        col_headers.append(short)
    col_headers.extend(["整体判断\nPASS/FAIL", "是否正确\n✅/❌", "模型输出理由", "备注"])

    # Row 4: 完整列头（提示行，可隐藏）
    for col, h in enumerate(col_headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

    # Row 4: 检查点完整文本（提示行）
    for i, cp in enumerate(checkpoints):
        c = ws.cell(row=4, column=cp_start + i, value=cp.name_cn)
        c.alignment = TOP_WRAP
        c.border = THIN_BORDER
        c.font = Font(size=8, color="666666")

    # ── 数据行 ──
    # 收集所有 unique cases (by case_id + sample_type)，按 case_id 排序
    all_keys = set()
    case_info: dict[str, dict] = {}
    for model, results in dim_data.items():
        for r in results:
            key = r["key"]
            all_keys.add(key)
            if key not in case_info:
                case_info[key] = r

    sorted_keys = sorted(all_keys, key=lambda k: (
        case_info[k]["case_id"],
        0 if case_info[k]["sample_type"] == "正例" else 1,
    ))

    row_idx = 5
    for key in sorted_keys:
        info = case_info[key]
        for model in models:
            # 找该模型对此 key 的结果
            model_results = dim_data.get(model, [])
            mr = next((r for r in model_results if r["key"] == key), None)

            # A-G: 案例信息
            ws.cell(row=row_idx, column=1, value=info["case_id"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=info["word"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=info["category"]).border = THIN_BORDER
            c = ws.cell(row=row_idx, column=4, value=info["sample_type"])
            c.border = THIN_BORDER
            c.alignment = CENTER
            c = ws.cell(row=row_idx, column=5, value=info.get("issue_desc", ""))
            c.border = THIN_BORDER
            c = ws.cell(row=row_idx, column=6, value=info["expected"])
            c.border = THIN_BORDER
            c.alignment = CENTER
            c = ws.cell(row=row_idx, column=7, value=model)
            c.border = THIN_BORDER
            c.alignment = CENTER

            if mr:
                # H~: 各检查点 Yes/No/NA
                cp_results = mr.get("checkpoints", {})
                for i, cp in enumerate(checkpoints):
                    val = cp_results.get(cp.id, "")
                    c = ws.cell(row=row_idx, column=cp_start + i, value=val)
                    c.alignment = CENTER
                    c.border = THIN_BORDER
                    if val.upper() == "YES":
                        c.font = YES_FONT
                    elif val.upper() == "NO":
                        c.font = NO_FONT
                        c.fill = FAIL_FILL

                # 整体判断
                judgment = mr.get("model_judgment", "")
                c = ws.cell(row=row_idx, column=summary_start, value=judgment)
                c.alignment = CENTER
                c.border = THIN_BORDER
                c.fill = PASS_FILL if judgment == "PASS" else FAIL_FILL

                # 是否正确
                is_correct = mr.get("is_correct", False)
                c = ws.cell(row=row_idx, column=summary_start + 1,
                            value="✅" if is_correct else "❌")
                c.alignment = CENTER
                c.border = THIN_BORDER

                # 模型输出理由
                reason = mr.get("reason", "")
                if not reason:
                    reason = mr.get("raw_output", "")[:300]
                c = ws.cell(row=row_idx, column=summary_start + 2, value=reason[:300])
                c.border = THIN_BORDER
                c.alignment = TOP_WRAP
            else:
                # 该模型没有此案例的结果
                for col in range(cp_start, summary_start + 3):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER

            # 备注列
            ws.cell(row=row_idx, column=summary_start + 3).border = THIN_BORDER

            row_idx += 1

    # ── 列宽 ──
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    # 检查点列
    for i in range(len(checkpoints)):
        col_letter = openpyxl.utils.get_column_letter(cp_start + i)
        ws.column_dimensions[col_letter].width = 10
    # 汇总列
    for i in range(4):
        col_letter = openpyxl.utils.get_column_letter(summary_start + i)
        ws.column_dimensions[col_letter].width = 12 if i < 2 else 40


# ── 汇总 sheet ─────────────────────────────────────────
def _write_summary_sheet(
    ws, all_data: dict, models: list[str], dims: list[str],
) -> None:
    row = 1

    def _write_table(title: str, metric_fn) -> int:
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13)
        row += 1
        headers = ["模型"] + [DIMENSION_CN[d] for d in dims] + ["平均/总体"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER
        row += 1

        for model in models:
            ws.cell(row=row, column=1, value=model).border = THIN_BORDER
            values = []
            for col, dim in enumerate(dims, 2):
                results = all_data.get(dim, {}).get(model, [])
                m = compute_metrics(results)
                val = metric_fn(m)
                values.append(val)
                c = ws.cell(row=row, column=col, value=f"{val*100:.1f}%")
                c.alignment = CENTER
                c.border = THIN_BORDER
                if metric_fn == _fpr and val > 0.1:
                    c.fill = FAIL_FILL
            avg = sum(values) / len(values) if values else 0
            c = ws.cell(row=row, column=len(dims) + 2, value=f"{avg*100:.1f}%")
            c.alignment = CENTER
            c.border = THIN_BORDER
            c.font = Font(bold=True)
            row += 1
        row += 2
        return row

    _acc = lambda m: m["accuracy"]
    _fpr = lambda m: m["fpr"]
    _fnr = lambda m: m["fnr"]

    _write_table("表1: 模型 × 维度 准确率", _acc)
    _write_table("表2: 假阳率 FPR（越低越好）⚠️ 最关键", _fpr)
    _write_table("表3: 假阴率 FNR（误杀率）", _fnr)

    # 推荐表
    ws.cell(row=row, column=1, value="表4: 按维度推荐模型").font = Font(bold=True, size=13)
    row += 1
    rec_headers = ["维度", "推荐模型", "准确率", "假阳率", "假阴率", "理由"]
    for col, h in enumerate(rec_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    for dim in dims:
        best_model, best_score, best_m = None, -1, {}
        for model in models:
            results = all_data.get(dim, {}).get(model, [])
            m = compute_metrics(results)
            score = m["accuracy"] - 1.0 * m["fpr"] - 0.5 * m["fnr"]
            if score > best_score:
                best_score, best_model, best_m = score, model, m

        ws.cell(row=row, column=1, value=DIMENSION_CN[dim]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=best_model or "N/A").border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{best_m.get('accuracy', 0)*100:.1f}%").border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"{best_m.get('fpr', 0)*100:.1f}%").border = THIN_BORDER
        ws.cell(row=row, column=5, value=f"{best_m.get('fnr', 0)*100:.1f}%").border = THIN_BORDER
        ws.cell(row=row, column=6, value="准确率最高且FPR最低").border = THIN_BORDER
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 16
    for col_letter in "BCDEFGHI":
        ws.column_dimensions[col_letter].width = 12


# ── 终端报告 ──────────────────────────────────────────
def print_report(all_data: dict[str, dict[str, list[dict]]]) -> None:
    models = sorted({m for dd in all_data.values() for m in dd})
    dims = [d for d in DIMENSION_CN if d in all_data]

    for title, metric_key in [
        ("表1: 准确率", "accuracy"),
        ("表2: 假阳率 FPR（越低越好）", "fpr"),
        ("表3: 假阴率 FNR", "fnr"),
    ]:
        print(f"\n{'='*70}\n{title}\n{'='*70}")
        header = f"{'模型':<12}" + "".join(f"{DIMENSION_CN[d]:>10}" for d in dims) + f"{'平均':>10}"
        print(header)
        print("-" * len(header))
        for model in models:
            vals = []
            row = f"{model:<12}"
            for dim in dims:
                m = compute_metrics(all_data.get(dim, {}).get(model, []))
                v = m[metric_key]
                vals.append(v)
                row += f"{v*100:>9.1f}%"
            avg = sum(vals) / len(vals) if vals else 0
            row += f"{avg*100:>9.1f}%"
            print(row)

    # 推荐
    print(f"\n{'='*70}\n表4: 按维度推荐\n{'='*70}")
    print(f"{'维度':<16} {'推荐':<12} {'准确率':>8} {'FPR':>8} {'FNR':>8}")
    print("-" * 60)
    for dim in dims:
        best_model, best_score, best_m = None, -1, {}
        for model in models:
            m = compute_metrics(all_data.get(dim, {}).get(model, []))
            score = m["accuracy"] - 1.0 * m["fpr"] - 0.5 * m["fnr"]
            if score > best_score:
                best_score, best_model, best_m = score, model, m
        print(
            f"{DIMENSION_CN[dim]:<16} {best_model or 'N/A':<12} "
            f"{best_m.get('accuracy', 0)*100:>7.1f}% "
            f"{best_m.get('fpr', 0)*100:>7.1f}% "
            f"{best_m.get('fnr', 0)*100:>7.1f}%"
        )


# ── xlsx 生成 ─────────────────────────────────────────
def generate_xlsx(all_data: dict[str, dict[str, list[dict]]]) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    models = [m for m in MODEL_ORDER if any(m in dd for dd in all_data.values())]
    # 加上不在预定义列表中的模型
    for dd in all_data.values():
        for m in dd:
            if m not in models:
                models.append(m)
    dims = [d for d in DIMENSION_CN if d in all_data]

    # 汇总 sheet
    ws_sum = wb.create_sheet(title="汇总", index=0)
    _write_summary_sheet(ws_sum, all_data, models, dims)

    # 每维度 detail sheet
    for dim in dims:
        _write_dimension_sheet(wb, dim, all_data.get(dim, {}), models)

    out_path = RESULTS_DIR / "质检模型评估结果.xlsx"
    wb.save(out_path)
    return out_path


def main() -> None:
    all_data = load_all_results()
    if not all_data:
        print("未找到评估结果文件。请先运行 python run_eval.py")
        return
    print_report(all_data)
    xlsx_path = generate_xlsx(all_data)
    print(f"\n评估表已生成: {xlsx_path}")


if __name__ == "__main__":
    main()
